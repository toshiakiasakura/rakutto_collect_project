from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import datetime 

path = "./test_data.xlsx"
shName1 = "患者情報入力シート"
shName2 = "患者プルタブシート" 

def main():
    wb = load_workbook(path)
    wbBase = load_workbook(path)
    ws1 = wb[shName1]
    wsBase = wbBase[shName1] 

    ws1.cell(1,1).value = "changed"
    print(ws1.cell(1,1).value, wsBase.cell(1,1).value)

def main1():
    path = "test_data_temp.xlsx"
    df = pd.read_excel(path, converters={"date":pd.Timestamp})
    print(df)

def main2():
    s = "2020/04/05" 
    d = datetime.datetime.strptime(s,"%Y/%m/%d")
    print(d)
if __name__ == "__main__":
    main2()
