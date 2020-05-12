import sys
import types 
import inspect
import datetime
import pandas as pd
import time 

import _utils 

from openpyxl import Workbook
from openpyxl import load_workbook

from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QLineEdit,\
            QFileDialog, QMessageBox,QPushButton, QComboBox,\
            QLabel, QVBoxLayout, QHBoxLayout, QInputDialog,\
            QGridLayout, QSpacerItem, QSizePolicy, QScrollArea
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QIcon, QGuiApplication, QColor




path = "./test_data.xlsx"
shName1 = "患者情報入力シート"
shName2 = "患者プルタブシート" 

rowValue       = "行番号" 
columnValue    = "列番号"  
columnItemName = "項目名"
preValue   = "変更前の値"
postValue  = "変更後の値" 



class Ui_scrollArea(_utils.basicUtils):
    def __init__(self):
        super().__init__()
        self.row = 3
        self.baseRow = 2
        self.baseCol = 1 
        self.colTypeRefRow = 1 
        self.colTypeRefLis = [self.tpNumeric, self.tpDate , self.tpDropDown]
        self.filter = []
        self.questionTitle = "confirmation"
        self.errorOccurFlag = False
        now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        
        self.path2Diff = path[:-5] + f"_{now}_差分_未.xlsx" 
        self.path2DiffFin = path[:-5] + f"_{now}_差分_済.xlsx"

        self.index = 1

    def initialize(self, scrollArea):
        self.setupGUI(scrollArea)
        self.readAllValuesFromSheet()

    def readData(self, path):
        self.wb = load_workbook(path)
        self.ws1 = self.wb[shName1]
        self.ws2 = self.wb[shName2] 

        self.maxRow = self.ws1.max_row
        self.maxRow2 = self.ws2.max_row
        self.maxBaseRow = self.baseRow
        self.maxColumn = self.ws1.max_column
        self.maxColumn2 = self.ws2.max_column

        # set colNameDic 
        self.colNameDic = {}
        self.colNameDic2 = {}
        self.indNameDic = {}
        self.colTypeDic = {}
        for col in range(1,self.maxColumn + 1 ): 
            colName = self.readOneValue(self.baseRow, col) 
            if colName == self.nonValue or col in self.filter:
                continue
            if colName in self.colNameDic.keys():
                raise Exception(f"{shName1} には、複数の同じ名称のカラムがあります。") 
            self.colNameDic[colName] = col 

        self.maxBaseRow = self.getMaxRow()

        for k,v in self.colNameDic.items():
            self.indNameDic[v] = k 

        for col in range(1, self.maxColumn2 + 1) :
            v = self.ws2.cell(self.baseRow, col).value
            v = self.convert2Str(v) 
            if v in self.colNameDic.keys() :
                if v in self.colNameDic2.keys():
                    raise Exception(f"{shName2} には、複数の同じ名称のカラムがあります。")
                self.colNameDic2[v] = col

                tp = self.ws2.cell(self.colTypeRefRow, col).value
                tp = self.convert2Str(tp)
                if tp in self.colTypeRefLis:
                    self.colTypeDic[v] = tp

        self.baseColValue= self.ws1.cell(self.baseRow, self.baseCol).value
        self.diffDic = {rowValue:[], columnValue:[], self.baseColValue:[], 
                columnItemName:[], preValue:[], postValue:[] }

    def setupGUI(self,scrollArea):
        scrollArea.setObjectName("scrollArea")
        mainWidth = 1000
        mainHeight = 600
        scrollArea.resize(mainWidth, mainHeight)
        scrollArea.setWidgetResizable(True)

        self.scrollAreaWidgetContents = QWidget()
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")

        self.scrollVerticalLayout = QVBoxLayout(self.scrollAreaWidgetContents)
        self.scrollVerticalLayout.setObjectName("scrollVerticalLayout")
        self.verticalLayoutScroll = QVBoxLayout()
        self.verticalLayoutScroll.setObjectName("verticalLayoutScroll")

        self.setupGridLayoutTop()
        self.setupScrollMiddle()
        self.setupGridLayoutBottom()
        scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.retranslateUi(scrollArea)

        QtCore.QMetaObject.connectSlotsByName(scrollArea)

    def changeRowRelated(self):
        reply = self.checkDiffMessage()
        if reply == QMessageBox.No:
            return()
        self.readAllValuesFromSheet()

    def readAllValuesFromSheet(self):
        print( inspect.currentframe().f_code.co_name) 

        num  = self.convertFromStr( self.lineRow.text(), self.tpNumeric  ) 
        if isinstance( num, int):
            #num += 1 
            self.row = num
        else:
            raise Exception("not int value is inputted")

        print(f"   row ref number : {num} , max row : {self.ws1.max_row}" )

        # change values 
        self.changeLineRef1( self.comboRef1.currentText() ) 
        self.changeLineRef2( self.comboRef2.currentText() ) 
        self.changeMacroValues()

        self.compareAllValues()

    def setupGridLayoutTop(self):
        self.gridLayoutTop = QGridLayout()
        self.gridLayoutTop.setObjectName("gridLayoutTop")

        # lineRow
        self.lineRow = QLineEdit(self.scrollAreaWidgetContents)
        sizePolicy = QSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineRow.sizePolicy().hasHeightForWidth())
        self.lineRow.setSizePolicy(sizePolicy)
        self.lineRow.setMinimumSize(QtCore.QSize(50, 0))

        self.lineRow.setObjectName("lineRow")
        self.lineRow.setText( str(self.row) )
        self.lineRow.editingFinished.connect(self.changeRowRelated)

        horizontalSpacerTop= QSpacerItem(50, 50, QSizePolicy.MinimumExpanding, QSizePolicy.Minimum)

        # comboRef1
        self.comboRef1 = QComboBox(self.scrollAreaWidgetContents)
        self.comboRef1.setObjectName("comboRef1")
        self.comboRef1.addItems( self.colNameDic.keys() ) 
        self.comboRef1.activated[str].connect(self.changeLineRef1)        

        # comboRef2 
        self.comboRef2 = QComboBox(self.scrollAreaWidgetContents)
        self.comboRef2.setObjectName("comboRef2")
        self.comboRef2.addItems( self.colNameDic.keys() ) 
        self.comboRef2.setCurrentIndex(1)
        self.comboRef2.activated[str].connect(self.changeLineRef2)        

        # lineRef1 
        self.lineRef1 = QLineEdit(self.scrollAreaWidgetContents)
        self.lineRef1.setObjectName("lineRef1")
        val =  self.getValueFromColumn( self.comboRef1.currentText() ) 
        self.lineRef1.setText(val)
        self.lineRef1.setEnabled(False) 

        # lineRef2
        self.lineRef2 = QLineEdit(self.scrollAreaWidgetContents)
        self.lineRef2.setObjectName("lineRef2")
        val =  self.getValueFromColumn( self.comboRef2.currentText() ) 
        self.lineRef2.setText(val)
        self.lineRef2.setEnabled(False) 

        # comboSearch 
        self.comboSearch = QComboBox(self.scrollAreaWidgetContents)
        self.comboSearch.addItems( self.colNameDic.keys() ) 
        self.comboSearch.setObjectName("comboSearch")
        self.comboSearch.activated[str].connect(self.changeLineSearch)        

        # lineSearch 
        self.lineSearch = QLineEdit(self.scrollAreaWidgetContents)
        self.lineSearch.setObjectName("lineSearch")
        val =  self.getValueFromColumn( self.comboSearch.currentText() ) 
        self.lineSearch.setText(val)

        # pushNewRow
        self.pushNewRow = QPushButton(self.scrollAreaWidgetContents)
        sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushNewRow.sizePolicy().hasHeightForWidth())
        self.pushNewRow.setSizePolicy(sizePolicy)
        self.pushNewRow.setObjectName("pushNewRow")
        self.pushNewRow.clicked.connect(self.addNewLine)
        # pushSearch
        self.pushSearch = QPushButton(self.scrollAreaWidgetContents)
        self.pushSearch.clicked.connect(self.findValue)
        self.pushSearch.setObjectName("pushSearch")


        # gridLayoutTop 
        self.gridLayoutTop.addWidget(self.lineRow, 0, 0, 1, 1)
        self.gridLayoutTop.addItem(horizontalSpacerTop, 0, 1, 1, 1)
        self.gridLayoutTop.addWidget(self.comboRef1, 0, 2, 1, 1)
        self.gridLayoutTop.addWidget(self.comboRef2, 1, 2, 1, 1)
        self.gridLayoutTop.addWidget(self.lineRef1, 0, 3, 1, 1)
        self.gridLayoutTop.addWidget(self.lineRef2, 1, 3, 1, 1)
        self.gridLayoutTop.addWidget(self.comboSearch, 0, 4, 1, 1)
        self.gridLayoutTop.addWidget(self.lineSearch, 0, 5, 1, 1)
        self.gridLayoutTop.addWidget(self.pushNewRow, 1, 0, 1, 1)
        self.gridLayoutTop.addWidget(self.pushSearch, 1, 4, 1, 2)
        self.verticalLayoutScroll.addLayout(self.gridLayoutTop)

    def setupScrollMiddle(self):
        self.scrollMiddle = QScrollArea(self.scrollAreaWidgetContents)
        self.scrollMiddle.setWidgetResizable(True)
        self.scrollMiddle.setObjectName("scrollMiddle")
        self.scrollMiddle.setMinimumSize(QtCore.QSize(500,300)) 
        self.scrollAreaWidgetContents_2 = QWidget()
        self.scrollAreaWidgetContents_2.setObjectName("scrollAreaWidgetContents_2")
        self.verticalLayout = QVBoxLayout(self.scrollAreaWidgetContents_2)
        self.verticalLayout.setObjectName("verticalLayout")

        self._horizontals = []
        self._labels = []
        self._comboBoxes = []
        self._lines = []
        self._horizontalSpacers = []

        for colName in self.colNameDic.keys():
            # check set new horizontal layout or not 
            self.setupMacros(colName)

        self.scrollMiddle.setWidget(self.scrollAreaWidgetContents_2)
        self.verticalLayoutScroll.addWidget(self.scrollMiddle)

    def setupMacros(self,colName):
        self.horizontalLayoutMacro = QHBoxLayout()
        self.horizontalLayoutMacro.setObjectName("horizontalLayoutMacro")

        # labelMacro
        self.labelMacro = QLabel(self.scrollAreaWidgetContents_2)
        self.labelMacro.setObjectName("labelMacro")
        self.labelMacro.setMinimumSize(QtCore.QSize(150,0)) 
        self.labelMacro.setText( colName )
        self.horizontalLayoutMacro.addWidget(self.labelMacro)

        # lineEdit setting. 
        self.lineMacro = QLineEdit(self.scrollAreaWidgetContents_2)
        self.lineMacro.setObjectName("lineMacro")
        self.lineMacro.setMinimumSize(QtCore.QSize(150,0)) 
        self.lineMacro.setStyleSheet(
                "QLineEdit { background-color : white; color:rgb(0,60,60)}")
        self.lineMacro._same = False 
        self.lineMacro.colName = colName
        def changeBackLine(obj:object):
            v = obj.text()
            colInd = self.colNameDic[obj.colName]
            origV  = self.readOneValue(self.row, colInd)
            tp = self.colTypeDic.get(obj.colName,None) 
            flag   = self.checkType( v,tp ) 
            if (flag == False) and (self.errorOccurFlag == True):
                print( inspect.currentframe().f_code.co_name) 
                print(v, origV, v==origV)
                if v != origV:
                    obj.setText(origV)
                    v = origV
                    self.typeErrorMessage(tp)
            
            b = self.compareValue( v , obj.colName ) 
            obj._same = b
            self.changeColor(obj, b, v)

        self.lineMacro.changeBackLine = types.MethodType(changeBackLine,self.lineMacro)
        self.lineMacro.editingFinished.connect( self.lineMacro.changeBackLine )

        v = self.getValueFromColumn( colName ) 
        self.lineMacro.setText( v ) 

        self.horizontalLayoutMacro.addWidget(self.lineMacro)

        # comboBox setting.
        self.comboBoxMacro = QComboBox(self.scrollAreaWidgetContents_2)
        self.lineMacro.combo = self.comboBoxMacro

        self.comboBoxMacro.setObjectName("comboBoxMacro")
        self.comboBoxMacro.setMinimumSize(QtCore.QSize(150,0)) 

        lis_ = self.getItemsFromSheet2(colName) 
        self.comboBoxMacro.addItems( lis_ ) 
        self.comboBoxMacro.colName = colName
        self.comboBoxMacro.defaultItems = lis_ 
        self.comboBoxMacro.line = self.lineMacro
        # initial color 
        self.comboBoxMacro.setStyleSheet("background-color: rgb(255,255,255);color:rgb(0,0,0);")

        def changeBackCombo(obj):
            v = obj.currentText()
            colInd = self.colNameDic[obj.colName]
            origV  = self.readOneValue(self.row, colInd)
            tp = self.colTypeDic.get(obj.colName,None)
            flag   = self.checkType(v, tp) 
            if flag == False and self.errorOccurFlag:
                self.setComboIndex(obj, origV)
                v= origV
                self.typeErrorMessage(tp)

            obj.line.setText(v) 

            b = self.compareValue( v, obj.colName) 
            obj.line._same = b
            self.changeColor(obj, b, v)

        self.comboBoxMacro.changeBackCombo = types.MethodType(changeBackCombo,self.comboBoxMacro)
        self.comboBoxMacro.activated[str].connect(self.comboBoxMacro.changeBackCombo)

        self.setComboIndex(self.comboBoxMacro,v) 

        self.horizontalLayoutMacro.addWidget(self.comboBoxMacro)

        # spacer  setting 
        horizontalSpacerMacro = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.horizontalLayoutMacro.addItem(horizontalSpacerMacro)
        self.verticalLayout.addLayout(self.horizontalLayoutMacro)

        # visible setting 
        tp = self.colTypeDic.get(colName, None)
        if tp == self.tpDropDown: 
            self.lineMacro.setVisible(False)
        else:
            self.comboBoxMacro.setVisible(False)

        # add each objects. 
        self._horizontals.append(self.horizontalLayoutMacro)
        self._labels.append(self.labelMacro)
        self._comboBoxes.append(self.comboBoxMacro)
        self._lines.append(self.lineMacro)
        self._horizontalSpacers.append(self.horizontalLayoutMacro)

    def setupGridLayoutBottom(self):
        self.gridLayoutBottom = QGridLayout()
        self.gridLayoutBottom.setObjectName("gridLayoutBottom")
        spacerItem2 = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)

        # label1
        self.labelSh1 = QLabel(self.scrollAreaWidgetContents)
        sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.labelSh1.sizePolicy().hasHeightForWidth())
        self.labelSh1.setSizePolicy(sizePolicy)
        self.labelSh1.setObjectName("labelSh1")

        # label2 
        self.labelSh2 = QLabel(self.scrollAreaWidgetContents)
        sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.labelSh2.sizePolicy().hasHeightForWidth())
        self.labelSh2.setSizePolicy(sizePolicy)
        self.labelSh2.setObjectName("labelSh2")

        # lineSh1
        self.lineSh1 = QLineEdit(self.scrollAreaWidgetContents)
        sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineSh1.sizePolicy().hasHeightForWidth())
        self.lineSh1.setSizePolicy(sizePolicy)
        self.lineSh1.setMinimumSize(QtCore.QSize(100, 0))
        self.lineSh1.setObjectName("lineSh1")
        self.lineSh1.setText(shName1)
        self.lineSh1.setEnabled(False) 

        # lineSh2 
        self.lineSh2 = QLineEdit(self.scrollAreaWidgetContents)
        sizePolicy = QSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineSh2.sizePolicy().hasHeightForWidth())
        self.lineSh2.setSizePolicy(sizePolicy)
        self.lineSh2.setObjectName("lineSh2")
        self.lineSh2.setText(shName2)
        self.lineSh2.setEnabled(False) 

        # pushWrite 
        self.pushWrite = QPushButton(self.scrollAreaWidgetContents)
        sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushWrite.sizePolicy().hasHeightForWidth())
        self.pushWrite.setSizePolicy(sizePolicy)
        self.pushWrite.setMinimumSize(QtCore.QSize(100, 0))
        self.pushWrite.setObjectName("pushWrite")

        self.pushWrite.clicked.connect(self.writeValues)

        # pushCancel
        self.pushCancel = QPushButton(self.scrollAreaWidgetContents)
        sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushCancel.sizePolicy().hasHeightForWidth())
        self.pushCancel.setSizePolicy(sizePolicy)
        self.pushCancel.setMinimumSize(QtCore.QSize(100, 0))
        self.pushCancel.setObjectName("pushCancel")

        self.pushCancel.clicked.connect(self.changeRowRelated)

        # pushSave
        self.pushSave= QPushButton(self.scrollAreaWidgetContents)
        self.pushSave.setMinimumSize(QtCore.QSize(100, 0))
        self.pushSave.setObjectName("pushSave")
        self.pushSave.clicked.connect(self.saveWorkBook)

        # pushQuit 
        self.pushQuit = QPushButton(self.scrollAreaWidgetContents)
        sizePolicy = QSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushQuit.sizePolicy().hasHeightForWidth())
        self.pushQuit.setSizePolicy(sizePolicy)
        self.pushQuit.setMinimumSize(QtCore.QSize(100, 0))
        self.pushQuit.setObjectName("pushQuit")

        self.pushQuit.clicked.connect(self.quitProcess)


        self.gridLayoutBottom.addItem(spacerItem2, 0, 2, 1, 1)
        self.gridLayoutBottom.addWidget(self.labelSh1, 0, 0, 1, 1)
        self.gridLayoutBottom.addWidget(self.labelSh2, 1, 0, 1, 1)
        self.gridLayoutBottom.addWidget(self.lineSh1, 0, 1, 1, 1)
        self.gridLayoutBottom.addWidget(self.lineSh2, 1, 1, 1, 1)
        self.gridLayoutBottom.addWidget(self.pushSave, 0, 5, 1, 1)
        self.gridLayoutBottom.addWidget(self.pushWrite, 1, 3, 1, 1)
        self.gridLayoutBottom.addWidget(self.pushCancel, 1, 4, 1, 1)
        self.gridLayoutBottom.addWidget(self.pushQuit, 1, 5, 1, 1)
        self.verticalLayoutScroll.addLayout(self.gridLayoutBottom)
        self.scrollVerticalLayout.addLayout(self.verticalLayoutScroll)

    def retranslateUi(self, scrollArea):
        _translate = QtCore.QCoreApplication.translate
        scrollArea.setWindowTitle(_translate("scrollArea", "ScrollArea"))
        self.pushNewRow.setText(_translate("scrollArea", "新規追加"))
        self.pushSearch.setText(_translate("scrollArea", "検索"))
        self.pushWrite.setText(_translate("scrollArea", "変更を追加"))
        self.labelSh1.setText(_translate("scrollArea", "入力先シート"))
        self.labelSh2.setText(_translate("scrollArea", "プルタブ参照シート"))
        self.pushSave.setText(_translate("scrollArea", "変更を保存")) 
        self.pushCancel.setText(_translate("scrollArea", "キャンセル"))
        self.pushQuit.setText(_translate("scrollArea", "終了"))

    def quitProcess(self):
        print( inspect.currentframe().f_code.co_name) 
        reply = self.preFinishProcess()
        if reply == QMessageBox.Yes:
            QApplication.instance().quit()
        else:
            pass 

    def closeEvent(self, event):
        # dose not work well.
        self.preFinishProcess()
        event.accept()

    def preFinishProcess(self):
        print( inspect.currentframe().f_code.co_name) 
        reply = self.checkDiffMessage()
        return(reply)

    def changeMacroValues(self):
        self.errorOccurFlag = False
        for i in range(len(self._lines) ) :
            colName = self._labels[i].text()
            v = self.getValueFromColumn( colName ) 

            # line setting 
            self._lines[i].setText(v)  
            # comboBox setting 
            combo = self._comboBoxes[i]
            combo.clear()
            combo.addItems( combo.defaultItems ) 
            self.setComboIndex(combo, v) 

        self.errorOccurFlag = True

    def setComboIndex(self, combo, v:str):
        index = combo.findText(v, QtCore.Qt.MatchFixedString)
        if index == - 1 :
            combo.insertItem(0,v)
            index = 0
        combo.setCurrentIndex(index) 

    def checkDiffExist(self):
        flag = False
        for i in range( len(self._lines) ) :
            if not self._lines[i]._same:
                flag = True
                return(flag)
        return(flag) 

    def checkDiffMessage(self,exp=None):
        if exp == None: 
            exp = "Yes を押すとこのページの変更が破棄されます。" 

        flag = self.checkDiffExist()
        reply = QMessageBox.Yes
        if flag :
            print(exp)
            reply = QMessageBox.question(self.scrollAreaWidgetContents, self.questionTitle,exp, 
                            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        return( reply ) 

    def typeErrorMessage(self,tp):
        print( inspect.currentframe().f_code.co_name) 
        exp = "想定外のエラーです。" 
        if tp == self.tpNumeric:
            exp = "整数を入力してください。"
        elif tp == self.tpDate:
            exp = "日付を入力してください。 \nex. 2020/04/01"

        QMessageBox.question(self.scrollAreaWidgetContents, self.questionTitle,exp, QMessageBox.Yes )
        

    def changeLineRef1(self, s:str) :
        v = self.getValueFromColumn(s)
        self.lineRef1.setText(v) 

    def changeLineRef2(self, s:str) :
        v = self.getValueFromColumn(s)
        self.lineRef2.setText(v) 

    def changeLineSearch(self, s:str):
        v = self.getValueFromColumn(s)
        self.lineSearch.setText(v)

    def changeColor(self,obj:object, compBool:bool, value:str):
        if compBool:
            if value == self.nonValue or value == self.empty: 
                obj.setStyleSheet(
                "background-color : rgb(230,230,255) ; color:rgb(0,60,60)")
            else:
                obj.setStyleSheet(
                "background-color : rgb(255,255,255); color:rgb(0,60,60)")
        else:
            obj.setStyleSheet(
            "background-color : yellow ; color:rgb(0,60,60)")
        

    def compareValue(self, v:str, colName):
        colInd = self.colNameDic[colName]
        origV = self.readOneValue(self.row, colInd) 
        if v == origV:
            return(True)
        else:
            return(False)

    def compareAllValues(self):
        self.errorOccurFlag = False
        for i in range(len(self._lines) ) :
            colName = self._lines[i].colName
            tp = self.colNameDic.get(colName, None)
            if tp == self.tpDropDown:
                combo = self._comboBoxes[i]
                combo.changeBackCombo()
            else:
                line = self._lines[i]
                line.changeBackLine()
        self.errorOccurFlag = True

    def findValue(self) :
        reply = self.checkDiffMessage()
        if reply == QMessageBox.No:
            return()
        self.readAllValuesFromSheet()            

        # find part.
        searchV = self.lineSearch.text()
        colName = self.comboSearch.currentText()
        colInd  = self.colNameDic[colName]
            
        colNameRef1 = self.comboRef1.currentText()
        colNameRef2 = self.comboRef2.currentText()
        colIndRef1 = self.colNameDic[ colNameRef1 ]
        colIndRef2 = self.colNameDic[ colNameRef2 ]

        findN = 0 
        findExp = f" 行 ,  {colNameRef1} ,  {colNameRef2} \n"
        for row in range(self.baseRow + 1 ,self.maxRow):
            v = self.readOneValue(row, colInd)
            if v == searchV: 
                self.lineRow.setText(str(row)) 
                findN += 1 
                valRef1 = self.readOneValue(row, colIndRef1) 
                valRef2 = self.readOneValue(row, colIndRef2) 
                findExp += f" {row} ,  {valRef1} ,  {valRef2} \n"

        if findN == 0:  
            exp1 = "検索した値は見つかりませんでした。\n"
        elif findN == 1:
            exp1 = "検索した値は見つかりました。\n"
            exp1 += findExp
        else:
            exp1 = "検索した値は、複数見つかりました。\n" 
            exp1 += findExp 

        QMessageBox.question(self.scrollAreaWidgetContents, 
                        self.questionTitle,exp1, QMessageBox.Yes )
        self.readAllValuesFromSheet()

    def readOneValue(self, row, col):
        v = self.ws1.cell(row,col).value
        v = self.convert2Str(v) 
        return(v) 

    def getItemsFromSheet2(self, colName):
        index = self.colNameDic2.get(colName, None)
        if index == None:
            return([])
        lis_ = []
        for row in range(self.baseRow + 1 , self.maxRow + 1 ):
            v = self.ws2.cell( row, index).value
            v = self.convert2Str(v) 
            if v == self.nonValue:
                return(lis_) 
            lis_.append(v)
        return(lis_)


    def getMaxRow(self): 
        for row in range(self.maxRow, self.baseRow,-1) :
            v = self.readOneValue(row, self.baseCol)
            if v == self.nonValue or v == self.empty:
                pass
            else:
                return(row)
        return(self.baseRow)

    def getMaxIndex(self):
        max_ = 1 
        for row in range(self.baseRow + 1, self.maxBaseRow + 1 ):
            v = self.readOneValue(row, self.baseCol) 
            v = self.convertFromStr(v, tp=self.tpNumeric)
            if isinstance(v, int):
                if max_ < v:
                    max_ = v 
        return(max_) 

    def addNewLine(self):
        reply = self.checkDiffMessage()
        if reply == QMessageBox.No:
            return()

        # add new data 
        self.maxBaseRow = self.getMaxRow() + 1 
        nextIndex = self.getMaxIndex() + 1 

        self.lineRow.setText( str(self.maxBaseRow) )
        self.row = self.maxBaseRow

        self.changeLineRef1( self.comboRef1.currentText() ) 
        self.changeLineRef2( self.comboRef2.currentText() ) 
        self.changeMacroValues()
        print( inspect.currentframe().f_code.co_name) 
        print(nextIndex)
        _translate = QtCore.QCoreApplication.translate
        # TO DO : If self.baseCol is changed, 
        #         this part does not correctly catch the behavior.  
        self._lines[0].setText(str(nextIndex) ) 
        self.compareAllValues()
        self._lines[0].setText(str(nextIndex)) 

    def getAllItemsCombo(self,combo):
        lis_ = []
        for i in range(combo.count()):
            lis_.append( combo.itemText(i) ) 
        return( lis_) 

    def getValueFromColumn(self,colName):
        colInd = self.colNameDic.get(colName ,None) 
        if colInd == None:
            print( inspect.currentframe().f_code.co_name) 
            print("Error there is no column name") 
            v = "eeeeerrrr"
        else:
            v = self.readOneValue( self.row, colInd) 
        return(v) 

    def writeValues(self):
        flag = self.checkDiffExist()
        if flag:
            print( inspect.currentframe().f_code.co_name) 
            self.addDiff2Dic()

        for i in range(len(self._lines) ) :
            colName = self._labels[i].text()
            colInd  = self.colNameDic[colName]
            tp      = self.colTypeDic.get(colName, None)
            v       = self._lines[i].text()
            v       = self.convertFromStr(v, tp)
            self.ws1.cell(self.row, colInd).value = v 
        self.readAllValuesFromSheet()
        self.compareAllValues()


    def addDiff2Dic(self):
        for i in range( len(self._lines) ) :
            line = self._lines[i]
            if not line._same:
                colInd = self.colNameDic[line.colName]
                baseV = self.readOneValue(self.row, self.baseCol)
                origV = self.readOneValue(self.row, colInd) 
                
                tp = self.colTypeDic.get(line.colName, None)
                baseV = self.convertFromStr( baseV , tp)
                origV = self.convertFromStr( origV , tp)
                newV  = self.convertFromStr( line.text(), tp) 

                self.diffDic[rowValue].append(self.row)
                self.diffDic[columnValue].append(colInd)
                self.diffDic[self.baseColValue].append(baseV)
                self.diffDic[columnItemName].append(line.colName)
                self.diffDic[preValue].append(origV)
                self.diffDic[postValue].append(newV)

        print(self.diffDic)
        with pd.ExcelWriter( self.path2Diff, engine="openpyxl", mode="wa", 
                datetime_format='yyyy/mm/dd') as writer:
            df = pd.DataFrame(self.diffDic) 
            df.to_excel(writer, sheet_name="差分", index=True)

    def saveWorkBook(self):
        st = time.perf_counter()

        pathOutput = path[:-5] + "_temp.xlsx"
        self.wb.save(pathOutput) 

        en = time.perf_counter()
        t = en- st 
        print(t)

def main():
    app = QApplication(sys.argv)

    scrollArea = QScrollArea()
    ui = Ui_scrollArea()
    ui.readData(path)
    ui.initialize(scrollArea)
    scrollArea.show()

    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
