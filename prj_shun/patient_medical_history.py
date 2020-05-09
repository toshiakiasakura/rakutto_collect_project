import openpyxl
import numpy as np
import pandas as pd

PATH = "/Users/shun/Desktop/nCoV_survey/nCoV_survey2004020_もっと修正案200504-2.xlsx"

def main():
    wb = openpyxl.load_workbook(PATH)

    #シート場所
    ws1 = wb.worksheets[0]

    #患者情報
    patient_id = "患者ID"
    patient_name = "患者氏名"

    patient_id_v = ws1.cell(row=4,column=38).value
    patient_name_v = ws1.cell(row=19,column=9).value

    #元データの疾患名, 列位置
    preg ='妊娠'
    smok = '喫煙'
    diab = '糖尿病'
    resp_dis = '呼吸器疾患（喘息・COPD・その他）'
    rena_dis = '腎疾患'
    live_dis = '肝疾患'
    hear_dis = '心疾患'
    nerv_dis = '神経筋疾患'
    bloo_dis = '血液疾患（貧血等）'
    immu = '免疫不全（HIV、免疫抑制剤使用含む）'
    canc = '悪性腫瘍（がん）'

    mhcp = medical_history_column_position = 4

    mark='■'
    no_mark="□"
    k1cp = key_1_column_position = 22
    k2cp = key_2_column_position = 25

    #疾患名出力表のカラム名変更
    preg_out ='妊娠'
    smok_out = '喫煙'
    diab_out = '糖尿病'
    resp_dis_out = '呼吸器疾患'
    rena_dis_out = '腎疾患'
    live_dis_out = '肝疾患'
    hear_dis_out = '心疾患'
    nerv_dis_out = '神経筋疾患'
    bloo_dis_out = '血液疾患'
    immu_out = '免疫不全'
    canc_out = '悪性腫瘍'

    #詳細情報1
    preg_week = '妊娠週数'
    smok_start_age = '喫煙：歳から'
    smok_per_day = '喫煙：本／日'
    dialysis = '（腎透析）'

    pwcp = preg_week_column_position = 31
    ssacp = smok_start_age_column_position = 29
    spdcp = smok_per_day_column_position = 35
    dk1cp = dialysis_key_1_column_position = 36
    dk2cp = dialysis_key_2_column_position = 40

    #詳細情報2
    resp_dis_detail = "呼吸器疾患詳細"
    live_dis_detail = "肝疾患詳細"
    hear_dis_detail = "心疾患詳細"
    nerv_dis_detail = "神経筋疾患詳細"
    bloo_dis_detail = "血液疾患詳細"
    immu_detail = "免疫不全詳細"
    canc_detail = "悪性腫瘍詳細"

    mhdcp = medical_history_column_position =33

    #その他疾患
    other_dis = "その他疾患"
    odcp = other_dis_column_position = 8

    #出力表のカラム順番変更
    out_order_list = [
        preg,
        preg_week,
        smok,
        smok_start_age,
        smok_per_day,
        diab,
        resp_dis,
        rena_dis,
        dialysis,
        live_dis,
        hear_dis,
        nerv_dis,
        bloo_dis,
        immu,
        canc,
        other_dis,
    ]

    out_list = [
        preg_out,
        preg_week,
        smok_out,
        smok_start_age,
        smok_per_day,
        diab_out,
        resp_dis_out,
        rena_dis_out,
        dialysis,
        live_dis_out,
        hear_dis_out,
        nerv_dis_out,
        bloo_dis_out,
        immu_out,
        canc_out,
        other_dis
    ]

    def getRowIndex(ws, row_loc_num, rowName):
        rowIndex = False
        for i in range(1,ws.max_row+1):
            if ws.cell(i, row_loc_num)._value  == rowName:
                rowIndex = i
                return(rowIndex)
        raise Exception(f"{rowName} の名前が見つかりませんでした。")

    def getRowIndexDict(ws,row_loc_num,rowNames) :
        dic_ = {}
        for r in rowNames: 
            dic_[r] = getRowIndex(ws,row_loc_num,r) 
        return(dic_) 

    def c_value(r,c):
        return ws1.cell(row=r,column=c).value

    col_1 =[
        preg,
        smok,
        diab,
        resp_dis,
        rena_dis,
        live_dis,
        hear_dis,
        nerv_dis,
        bloo_dis,
        immu,
        canc
    ]

    dic_1 = getRowIndexDict(ws1,mhcp,col_1)

    yes_no_l = []
    for i in dic_1.values():
        v1 = c_value(i,k1cp)
        v2 = c_value(i,k2cp)
        if (v1 == mark) & (v2 == no_mark):
            x = c_value(i,k1cp+1)
        elif (v1 == no_mark) & (v2 == mark):
            x = c_value(i,k2cp+1)
        elif (v1 == no_mark) & (v2 == no_mark):
            x = "記入なし"
        else:
            x = "error"
        yes_no_l.append(x)

    dic_1_yn = {}
    for k,v in zip(col_1,yes_no_l):
        dic_1_yn[k]=v

    dic_2 = {
        preg_week:[c_value(dic_1[preg],pwcp)],
        smok_start_age:[c_value(dic_1[smok],ssacp)],
        smok_per_day:[c_value(dic_1[smok],spdcp)],
    }

    s1=c_value(dic_1[rena_dis],dk1cp)
    s2=c_value(dic_1[rena_dis],dk2cp)

    if (dic_1_yn[rena_dis]==c_value(dic_1[rena_dis],k1cp+1)) & (s1==no_mark) & (s2==no_mark): 
        x = "腎疾患なし"
    elif (dic_1_yn[rena_dis]==c_value(dic_1[rena_dis],k2cp+1)) & (s1==mark) & (s2==no_mark): 
        x = c_value(dic_1[rena_dis],dk1cp+1)
    elif (dic_1_yn[rena_dis]==c_value(dic_1[rena_dis],k2cp+1)) & (s1==no_mark) & (s2==mark): 
        x = c_value(dic_1[rena_dis],dk2cp+1)
    elif (dic_1_yn[rena_dis]==c_value(dic_1[rena_dis],k2cp+1)) & (s1==no_mark) & (s2==no_mark): 
        x = "透析記入なし"
    else:
        x = "error"

    dic_3 = {
        dialysis:x
    }

    dic_4 = {
        resp_dis_detail:[c_value(dic_1[resp_dis],mhdcp)],
        live_dis_detail:[c_value(dic_1[live_dis],mhdcp)],
        hear_dis_detail:[c_value(dic_1[hear_dis],mhdcp)],
        nerv_dis_detail:[c_value(dic_1[nerv_dis],mhdcp)],
        bloo_dis_detail:[c_value(dic_1[bloo_dis],mhdcp)],
        immu_detail:[c_value(dic_1[immu],mhdcp)],
        canc_detail:[c_value(dic_1[canc],mhdcp)],
    }

    dic_5 = {
        other_dis:[c_value(max(dic_1.values())+1,odcp)]
    }

    dic_merge = dic_1_yn
    dic_merge.update(dic_2)
    dic_merge.update(dic_3)
    dic_merge.update(dic_5)

    collect_data_list = []
    for i in out_order_list:
        collect_data_list.append(dic_merge[i])

    dic_all = {}
    for k,v in zip(out_list,collect_data_list):
        dic_all[k]=v

    dic_all.update(dic_4)

    dic_patient = {
        patient_id:patient_id_v,
        patient_name:patient_name_v,
    }

    dic_patient.update(dic_all)

    df = pd.DataFrame(dic_patient)

    pathOutput = PATH[:-5] + "_Patient_Medical_History.xlsx"
    df.to_excel(pathOutput)

if __name__ == "__main__":
	main()