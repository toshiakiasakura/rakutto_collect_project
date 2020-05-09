import openpyxl
import numpy as np
import pandas as pd

PATH = "./nCoV_survey2004020_もっと修正案200504-2.xlsx"

def main():
    wb=openpyxl.load_workbook(PATH)

    ws1=wb.worksheets[0]
    ws4=wb.worksheets[3]

    #患者情報
    patient_id = "患者ID"
    patient_name = "患者氏名"

    patient_id_v = ws1.cell(row=4,column=38).value
    patient_name_v = ws1.cell(row=19,column=9).value

    #元データのカラム名、カラム位置、マーク文字、マークサーチ範囲
    ct_num="接触者"
    name="氏名"
    yomi="よみがな"
    rel="続柄"
    age="年齢"
    sex="性別"
    ct_time="患者との"
    ud_dis="基礎"
    onset="観察期間内"
    adr="連絡先（電話番号、"
    other="備考（接触状況等）"

    cln=5

    mark='■'
    no_mark="□"
    cvr=3

    col_1=[ct_num,name,yomi,rel,age,adr,other]
    col_2=[sex, ud_dis, onset]
    col_3 = [ct_time]

    #出力表のカラム名変更
    ct_num_out="接触者番号"
    name_out="氏名"
    yomi_out="よみがな"
    rel_out="続柄（関係）"
    age_out="年齢"
    sex_out="性別"
    ct_time_out="患者との最終接触日"
    ud_dis_out="基礎疾患"
    onset_out="観察期間内の発症"
    adr_out="連絡先（電話番号、メールアドレス等）"
    other_out="備考（接触状況等）"

    #出力表のカラム順番変更
    out_order_list = [
        ct_num,
        name,
        yomi,
        rel,
        age,
        sex,
        ct_time,
        ud_dis,
        onset,
        adr,
        other,
    ]

    out_list = [
        ct_num_out,
        name_out,
        yomi_out,
        rel_out,
        age_out,
        sex_out,
        ct_time_out,
        ud_dis_out,
        onset_out,
        adr_out,
        other_out,
    ]

    def getColIndex(ws,col_loc_num,colName):
        colIndex = False
        for  i in range(1,ws.max_column+1):
            if ws.cell(col_loc_num, i)._value  == colName:
                colIndex = i
                return(colIndex)
        raise Exception(f"{colName} の名前が見つかりませんでした。")

    def getColIndexDict(ws,col_loc_num,colNames) :
        dic_ = {}
        for c in colNames: 
            dic_[c] = getColIndex(ws,col_loc_num,c) 
        return(dic_) 

    dic_1=getColIndexDict(ws4,cln,col_1)

    name_l=[]
    for i in range(cln+2,ws4.max_row-1):
        x=ws4.cell(row=i, column=dic_1[name]).value
        if x==None:
            break
        name_l.append(x)
    col_len=len(name_l)

    dic_l_1 = {}
    for k,v in dic_1.items():
        l = list()
        for i in range (cln+2,col_len+cln+2):
            l.append(ws4.cell(row=i, column=v).value)
        x = {k:l}
        dic_l_1.update(x)

    dic_2=getColIndexDict(ws4,cln,col_2)

    def seeSameRowItems(ws,row_n,start_col_n,cover_range):
        l = []
        for i in range(1,cover_range+1):
            l.append((ws.cell(row=row_n, column=start_col_n+i-1)).value)
        return l

    def getItems_SeeSameRowItems(ws,row_n,start_col_n,cover_range,item):
        x = seeSameRowItems(ws,row_n,start_col_n,cover_range)
        if (item in x):
            return item

    dic_l_2 = {}
    for k,v in dic_2.items():
        l = list()
        for i in range (cln+2,col_len+cln+2):
            v1 = ws4.cell(row=i, column=v).value
            v2 = getItems_SeeSameRowItems(ws4,i,v+1,cvr,mark) or getItems_SeeSameRowItems(ws4,i,v+1,cvr,no_mark)
            if (v1 == mark) & (v2 == no_mark):
                x = ws4.cell(row=i, column=v+1).value
            elif (v1 == no_mark) & (v2 == mark):
                n = seeSameRowItems(ws4,i,v+1,cvr).index(mark)
                x = ws4.cell(row=i, column=v+n+2).value
            elif (v1 == no_mark) & (v2 == no_mark):
                x = "入力なし"
            else:
                x = "error"
            l.append(x)
        x = {k:l}
        dic_l_2.update(x)

    dic_3 = getColIndexDict(ws4,cln,col_3)

    dic_l_3 = {}
    for k,v in dic_3.items():
        l = list()
        for i in range (cln+2,col_len+cln+2):
            a = ws4.cell(row=i, column=v).value
            b = ws4.cell(row=i, column=v+3).value
            c = ws4.cell(row=i, column=v+6).value

            x = str(a)+"/"+str(b)+"/"+str(c)
            y = pd.to_datetime(x)

            l.append(y)

        x = {k:l}
        dic_l_3.update(x)

    dic_merge = dic_l_1
    dic_merge.update(dic_l_2)
    dic_merge.update(dic_l_3)

    collect_data_list = []
    for i in out_order_list:
        collect_data_list.append(dic_merge[i])

    dic_all = {}
    for k,v in zip(out_list,collect_data_list):
        dic_all[k]=v

    patient_id_l = []
    for i in range (1,col_len+1):
        patient_id_l.append(patient_id_v)

    patient_name_l = []
    for i in range (1,col_len+1):
        patient_name_l.append(patient_name_v)

    dic_patient = {
        patient_id:patient_id_l,
        patient_name:patient_name_l,
    }

    dic_patient.update(dic_all)

    df = pd.DataFrame(dic_patient)

    pathOutput = PATH[:-5] + "_List_of_Contact_Person.xlsx"
    df.to_excel(pathOutput)
    
if __name__ == "__main__":
	main()
