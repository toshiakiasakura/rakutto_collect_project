
import pandas as pd
import numpy as np
import os
import datetime
import _utils 

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color,colors,  Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

#import warnings
#warnings.simplefilter('ignore')

import pathList as pL 

if os.name == "nt":
    import locale
    locale.setlocale(locale.LC_CTYPE, "Japanese_Japan.932")

shName = pL.testSheetName
pullName = pL.testSheetPull

def main(path):


    extendFormat(path)


def extendFormat(path,change=True):

    # read files 
    pullDown = pd.read_excel(path ,
            header = 1 ,sheet_name=pL.testSheetPull,
            encoding = "cp932")

    wb = load_workbook(filename = path)
    ws = wb[shName]

    max_ = ws.max_row
    for c in pullDown.columns:
        
        if "Unnamed" in c:
            continue
        if pullDown[c].dropna().shape[0] == 0:
            continue
        colIndex = _utils.getColIndex(ws,c) 
        colStr = get_column_letter(colIndex)
        #print(c,colStr, pullDown[c].dropna().shape[0])
        ws  = setDropDown(ws,pullDown =  pullDown, 
                colName = c, colStr = colStr,max_= max_+1000)

    if change:
        pathOutput = path[:-5] + "_output.xlsx"
    else:
        pathOutput = path
    wb.save(pathOutput)


def setDropDown(ws,pullDown=None,colName=None,colStr=None,fm1=None,max_=1048576): 
    if fm1 == None:
        s = ",".join(pullDown[colName].dropna().unique())
        fm1 =  f'"{s}"'
    if s == "":
        return(ws)
    dv = DataValidation(type="list", formula1=fm1)
    dv.ranges = f'{colStr}3:{colStr}{max_}'
    ws.add_data_validation(dv)
    return(ws)

if __name__ == "__main__":
    import pathList as pL
    main(pL.newStylePatient)

