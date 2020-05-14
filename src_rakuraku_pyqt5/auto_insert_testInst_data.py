import pandas as pd
import numpy as np
import datetime
import os 
import glob
import copy

import _utils
import pathList as pL

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color,colors,Border,Side,  Alignment, PatternFill
blueFill = PatternFill("solid",bgColor="d5e6f3",fgColor ="d5e6f3" )
redFill = PatternFill("solid",bgColor="fb0007",fgColor ="fb0007" )
skinFill = PatternFill("solid", bgColor="fceac5", fgColor="fceac5")

if os.name == "nt":
    import locale
    locale.setlocale(locale.LC_CTYPE, "Japanese_Japan.932")



key1 = '検査機関\n（入力必須）'
key2 = '検査機関内番号'
key3 = "元ファイル"
No = "番号" 
item = "項目名" 
testName = "検査データの値" 
testInstName = "検査機関データの値" 

def main(dirInst, path):
    pathInsts  = glob.glob(dir_ + "*")

    dfs = []
    for p in pathInsts:
        df = pd.read_excel(p , encoding = "cp932", header = 1)
        df = df.dropna(axis = 0 ,how= "all")
        dfs.append(df)
    path  =  pL.testCleaned
    test   =  pd.read_excel(path, encoding = "cp932", header = 1)


    autoInsert(dfs,test,path)


def autoInsert(dfs, test,pathInsts, path):

    dfInst, errorMsg = mergeDFs(pathInsts,dfs,key1,key2) 
    if errorMsg:
        return(errorMsg)

    # check duplicated value exists or not 
    errorMsg = checkDuplication(dfInst, key1,key2)
    if errorMsg:
        return(errorMsg) 

    testNew, dfRest, dfExist  = preProcessing(test,dfInst, key1, key2)

    # add data with openpyxl 
    shName = pL.testSheetName
    wb = load_workbook(filename = path) #read_only=False, keep_vba=True)
    wb = deleteErrorSheet(wb) 
    ws = wb[shName]

    max_row = ws.max_row
    n = dfRest.shape[0]
    
    colIndDic = _utils.getColIndexDict(ws, pL.insertCols ) 

    NoIndex = _utils.getColIndex(ws, No)
    errorDic = { No:[], key1:[], key2:[], key3:[], item :[],
            testName:[], testInstName:[] }  


    # add data to already existing data frame
    ind1 = colIndDic[key1]
    ind2 = colIndDic[key2]

    for row in range(1, max_row+ 1 ):
        v1 = ws.cell(row,ind1)._value
        v2 = ws.cell(row,ind2)._value
        if not (v1 and v2):
            continue

        cond = ( dfExist[key1] == v1) & ( dfExist[key2] == v2) 
        dfM = dfExist.loc[cond] 
        if dfM.shape[0] > 1 :
            raise Error(f"{key1} and {key2} multiple data are found") 
        elif dfM.shape[0] == 1:
            for c, colIndex in colIndDic.items():
                # data that will be inserted exist
                insertValue = dfM.iloc[0][c]
                if insertValue != insertValue: 
                    continue

                cell = ws.cell(row,colIndex)
                value = cell.value
                if value == None: 
                    cell.value = insertValue
                    cell.fill = skinFill
                    if cell.has_style:
                        cell_base = ws.cell(row=3,column = colIndex) 
                        cell._style = copy.copy(cell_base._style)
                        cell.font   = copy.copy(cell_base.font)
                else:
                    if isinstance(value,datetime.date):
                        value = value.date()
                    if value != insertValue:
                        cell.fill = redFill
                        # TO DO : error data should be appeared in another sheet.
                        errorDic[No].append( ws.cell(row,NoIndex).value )
                        errorDic[key1].append( ws.cell(row,colIndDic[key1]).value )
                        errorDic[key2].append( ws.cell(row,colIndDic[key2]).value )
                        errorDic[key3].append( dfM.iloc[0][key3].split("/")[-1]  )
                        errorDic[item].append( c ) 
                        errorDic[testName].append( value )
                        errorDic[testInstName].append( insertValue )
                    else:
                        cell.fill = skinFill



    for c, colIndex in colIndDic.items():
        for i in range(n):
            cell = ws.cell(row=max_row + i + 1 , column=colIndex)
            v = dfRest.iloc[i][c]

            if v == v:
                cell.value = v
            if cell.has_style:
                cell_base = ws.cell(row=3,column = colIndex) 
                cell._style = copy.copy(cell_base._style)
                cell.font   = copy.copy(cell_base.font)
            if v == v:
                cell.fill = skinFill

    ws.auto_filter.ref = 'A2:AA1048576' # cf. ws.dimensions, add auto_filter.
#    dir_ = _utils.getOutputDir()
#    pathOutput = dir_ + path.split("/")[-1][:-5] + "_追加済.xlsx"
    pathOutput = path[:-5] + "_追加済み.xlsx" 
    wb.save(pathOutput)


    pathError = path[:-5] + "_エラー.xlsx"
    with pd.ExcelWriter( pathError, engine="openpyxl", mode="w") as writer:
        errorDF = pd.DataFrame(errorDic) 
        errorDF.to_excel(writer, sheet_name = pL.testSheetError,index=False) 
    

def deleteErrorSheet(wb):
    if pL.testSheetError in wb.sheetnames:
        wb.remove(wb[pL.testSheetError]) 
#    wb.create_sheet(title = pL.testSheetError) 
    return(wb) 


def mergeDFs(pathInsts,dfs, key1, key2) :
    errorMsg = ""
    columns = np.hstack((dfs[0].columns,key3))  
    dfInst = pd.DataFrame(columns = columns)

    for p ,df in zip(pathInsts, dfs):
        print(p, df.shape)
        df[key3] = p
        condNan = df[key1].isnull()
        if condNan.sum() > 0 :
            pc = p.split("/")[-1]
            key1c = key1.replace("\n","")
            errorMsg += f"'{pc}'の'{key1c}'には欠損があります。\n"

        condNan = df[key2].isnull()
        if condNan.sum() > 0:
            pc = p.split("/")[-1]
            key2c = key2.replace("\n","")
            errorMsg += f"'{pc}'の'{key2c}'には欠損があります。\n"

        dfInst = pd.concat((dfInst,df))
    dfInst = dfInst.reset_index(drop=True)
    return(dfInst, errorMsg)


def checkDuplication(dfInst,key1,key2):
    # TO DO : make it easy to find the source files.
    errorMsg = ""
    for c in dfInst[ key1 ].unique():
        cond = dfInst[ key1 ] == c 
        dfM = dfInst.loc[cond,key2 ] 
        counts = dfM.value_counts() > 1
        if counts.sum() > 0 :
            errorMsg += f"'{c}' には重複している番号があります。\n"
    return(errorMsg)


def preProcessing(test,dfInst, key1,key2):
    for c in dfInst.columns:
        dfInst[c] = _utils.convertDatetime(dfInst, c) 

    # testNew is data frame that contains key1 and key2
    cond1 = test[key1].isnull() == False
    cond2 = test[key2].isnull() == False
    testNew = test.loc[cond1&cond2].reset_index()

    # extract test institution data which do not exist in test data.  
    dfRest = dfInst.copy()
    for i in testNew.index : 
        cond1 = testNew.loc[i,key1]  == dfRest[key1]
        cond2 = testNew.loc[i,key2]  == dfRest[key2]
        condRest = 1^(cond1&cond2)
        dfRest = dfRest.loc[condRest]
            
    dfExist = dfInst.drop(dfRest.index)

    print(dfInst.shape, dfRest.shape, dfExist.shape ) 
    print(test.shape, testNew.shape) 
    return(testNew, dfRest, dfExist) 


if __name__ == "__main__":
    # still not implemented 
    main(dirInst, path)


