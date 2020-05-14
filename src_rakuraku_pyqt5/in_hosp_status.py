# ======================================================================
# Project Name    : rakutto_collect_project
# File Name       : in_hosp_status.py
# Encoding        : utf-8
# Copyright © 2020 Toshiaki Asakura. All rights reserved.
# ====================================================================== #

import numpy as np
import pandas as pd
import datetime
import openpyxl
import os

import _utils
import pathList as pL

import warnings
warnings.simplefilter('ignore')

if os.name == "nt":
    import locale
    locale.setlocale(locale.LC_CTYPE, "Japanese_Japan.932")

program = "in_hosp_status.py"


today = datetime.datetime.now().strftime('%y%m%d')
dir_ = _utils.getOutputDir()
output1 = dir_ + f"医療圏-機関-患者-対応表-公開用_{today}.xlsx"
output2 = dir_ + f"医療圏-機関-患者-対応表_{today}.xlsx"


# for global variables of formatExcelStyle() 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color,colors,Border,Side,  Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

# set colors
darkGreyFill = PatternFill("solid",bgColor="A9A9A9",fgColor ="A9A9A9" )
baseAlign = Alignment(wrapText=True,horizontal="center", vertical="top")
alignCenter = Alignment(wrapText=True,horizontal="center", vertical="center")
borderType = 'thin'
border = Border(left=Side(border_style=borderType,
                           ),
                 right=Side(border_style=borderType,
                           ),
                 top=Side(border_style=borderType,
                          ),
                 bottom=Side(border_style=borderType,
                           ),
                 outline=Side(border_style=borderType),
               )
#borderType = 'thick'
#borderThick = Border(left=Side(border_style=borderType,
#                           ),
#                 right=Side(border_style=borderType,
#                           ),
#                 top=Side(border_style=borderType,
#                          ),
#                 bottom=Side(border_style=borderType,
#                           ),
#                 outline=Side(border_style=borderType),
#               )


hosp = "入院医療機関（現在）"
sex = "性別"
address = "居住地"
bed = "入院病床（現在）"
No = "№"
out = "退院日"
numHosp = "入院患者数"

physical = "身体状況（現在の症状）"
secondRegion = "二次医療圏" 
hospitalName = "病院名" 
numInfHosp = "感染症病床数" 

def main(path1, path2):
    try:
        df = pd.read_excel(path1,
                        sheet_name=pL.patientSheetName ,
                        header = 1,
                        encodeing="cp932"
                        )
        hospital = pd.read_excel(path2,encoding='cp932',sheet_name='Sheet1')
    except:
        errorMsg = "データが読み込めませんでした。\n"
        import traceback
        errorMsg  +=  traceback.format_exc()
        return(errorMsg)


    errorMsg = getErrorMsg(df,hospital)
    if errorMsg == False:
        try:
            mergeProcessFiles(df,hospital)
            #formatExcelStyle()
            _utils.createErrorCheckFile(path1,program = program) 
            _utils.createErrorCheckFile(path2,program = program) 
        except:
            errorMsg  = "予期せぬエラーが発生しました。\n"
            import traceback
            errorMsg  +=  traceback.format_exc()


    print(errorMsg)    
    return(errorMsg)


def getErrorMsg(df,hospital):
    return(False)



def mergeProcessFiles(df,hospital):
#    hospital = hospital.replace({".+_(.+)":r"\1"},regex=True)  
#    df[hosp] = df[hosp].replace({".+_(.+)":r"\1"},regex=True)  
    hospitalInf = hospital[[secondRegion,hospitalName,numInfHosp]]


    df["count"] = 1 
    df['備考'] = df['転帰'].replace(np.nan, '') \
              + ' '+ df['居住地：備考'].replace(np.nan, '') \
              + ' '+ df['職業：備考'].replace(np.nan, '')\
              + ' '+ df['発症日：備考'].replace(np.nan, '')\
              + ' '+ df['濃厚接触者の観察状況：備考'].replace(np.nan, '')\
              + ' '+ df['入院：備考'].replace(np.nan, '')\
              + ' '+ df['転院：備考'].replace(np.nan, '')\


    #   get 入院状況
    hospFlag = '入院患者' 
    in_ = "入院中"
    df[hospFlag] = np.nan  
    cond1 = df[physical].str.contains("陰性確認").replace(np.nan,False) == False
    cond2 = df[physical].str.contains("死亡").replace(np.nan,False) == False
    df.loc[ cond1 & cond2 , hospFlag ] =  in_
    df[numHosp] = 0
    df.loc[ cond1 & cond2  ,numHosp] = 1 

    #   感染症病床のみ
    kind  = "入院病床（現在）"
    condInf  = df[kind].str.contains( "感染症病床").replace(np.nan,False) 
    dfInf = df.loc[condInf] 

    table1 = mergeTwoTable(dfInf, hospitalInf, hospFlag) 
    #   その他も含めたもの。
    table2 = mergeTwoTable(df,hospital, kind)

    with pd.ExcelWriter(output1, engine = 'openpyxl' 
            , mode='wa',datetime_format='mm/dd') as writer: 
        table1.to_excel(writer, index=False , sheet_name="第一種・第二種感染症指定医療機関_病床数")
        table2.to_excel(writer, index=False , sheet_name="その他医療機関_病床数")

    # 入院状況
    dfDetail = df[['№','年代','性別','居住地２','入院日','退院日','備考', hosp ,kind]]
    dfMerged = pd.merge(hospital, dfDetail, left_on=hospitalName, right_on = hosp, how='outer')
    dfMerged = dfMerged.sort_values(by=[secondRegion,hospitalName],ascending=True,na_position="last") 

    condInf = dfMerged[kind].str.contains( "感染症病床").replace(np.nan, False) 
    dfMergedInf = dfMerged.loc[condInf]
    dfMergedOther = dfMerged.loc[1^condInf]

    with pd.ExcelWriter(output1, engine = 'openpyxl' 
            , mode='wa',datetime_format='mm/dd') as writer: 
        table1.to_excel(writer, index=False , sheet_name="第一種・第二種感染症指定医療機関_病床数")
        table2.to_excel(writer, index=False , sheet_name="その他医療機関_病床数")
        dfMergedInf.to_excel(writer, index=False , sheet_name="第一種・第二種感染症指定医療機関_入院状況")
        dfMergedOther.to_excel(writer, index=False , sheet_name="の他医療機関_入院状況")

def mergeTwoTable(df_, hospital_, col) :
    tableHosp = df_.pivot_table(values="count", aggfunc=np.sum, 
                index = hosp, columns = col)
    tableHosp = tableHosp.reset_index()

    table1 = pd.merge(hospital_ , tableHosp, left_on = hospitalName , right_on = hosp, how='outer')
    table1[hospitalName] = np.where( table1[hospitalName].isnull()^1, table1[hospitalName], table1[hosp])
    table1= table1.sort_values(by=[secondRegion,hospitalName],ascending=True,na_position="last") 
    return(table1)


def formatExcelStyle():
    # ブックを取得
    book = openpyxl.load_workbook(output1)

    # シートを取得
    sheet1 = book['第一種・第二種感染症指定医療機関_病床数']
    sheet2 = book['その他医療機関_病床数']
    sheet3 = book['第一種・第二種感染症指定医療機関_入院状況']
    sheet4 = book['その他医療機関_入院状況']


    #行挿入
    sheet1.insert_rows(1)
    sheet2.insert_rows(1)
    sheet3.insert_rows(1)
    sheet3.insert_rows(2)
    sheet3.insert_rows(3)
    sheet3.insert_rows(4)
    sheet3.insert_rows(5)
    sheet4.insert_rows(1)
    sheet4.insert_rows(2)


    #セル結合
    sheet1.merge_cells('A1:D1')

    sheet2.merge_cells('A1:A2')
    sheet2["A1"] = "第二次\n医療圏"

    sheet2.merge_cells('B1:B2')
    sheet2["B1"] = "医療機関名"

    sheet2.merge_cells('C1:E1')
    sheet2["C1"] = "確保病床数"
    sheet2["C2"] = "感染症\n病床数"

    sheet2.merge_cells('F1:H1')
    sheet2["F1"] = "入院患者数"
    sheet2["F2"] = "感染症\n病床"
    sheet2["G2"] = "その他"
    sheet2["H2"] = "計"


    sheet3.merge_cells('A5:A6')
    sheet3["A5"] = "第二次\n医療圏"

    sheet3.merge_cells('B5:B6')
    sheet3["B5"] = "医療機関名"

    sheet3.merge_cells('C5:C6')
    sheet3["C5"] = '指定\n病床数'

    sheet3.merge_cells('D5:D6')
    sheet3["D5"] = '入院\n患者数'

    sheet3.merge_cells('E5:K5')
    sheet3["E5"] = '内　　　訳'
    sheet3["E6"] = 'No.'
    sheet3["F6"] = '年齢'
    sheet3["G6"] = '性別'
    sheet3["H6"] = '居住地'
    sheet3["I6"] = '入院月日'
    sheet3["J6"] = '退院月日'
    sheet3["K6"] = '備　　　考'

    sheet3['K2'] = '取扱注意'

    sheet3.merge_cells('A3:K3')
    sheet3.merge_cells('B4:J4')
    sheet3['B4'] = '【第一種・第二種感染症指定医療機関】'



    sheet4.merge_cells('A2:A3')
    sheet4["A2"] = "第二次\n医療圏"

    sheet4.merge_cells('B2:B3')
    sheet4["B2"] = "医療機関名"

    sheet4.merge_cells('C2:C3')
    sheet4["C2"] = '指定\n病床数'

    sheet4.merge_cells('D2:D3')
    sheet4["D2"] = '入院\n患者数'

    sheet4.merge_cells('E2:K2')
    sheet4["E2"] = '内　　　訳'
    sheet4["E3"] = 'No.'
    sheet4["F3"] = '年齢'
    sheet4["G3"] = '性別'
    sheet4["H3"] = '居住地'
    sheet4["I3"] = '入院月日'
    sheet4["J3"] = '退院月日'
    sheet4["K3"] = '備　　　考'

    sheet4.merge_cells('B1:J1')
    sheet4['B1'] = '【その他の医療機関】'



    #列幅
    widthDic = {"A":25, "B":50, "C" : 25, "D":25}
    for k,v in widthDic.items():
        sheet1.column_dimensions[k].width = v 

    widthDic = {"A":20, "B":50 }
    for k,v in widthDic.items():
        sheet2.column_dimensions[k].width = v 

    for c in "CDEFGH":
        sheet2.column_dimensions[c].width = 15

    widthDic = {"A":15,
                "B":50,
                "C":10,
                "D":10,
                "E":19,
                "F":12,
                "G":10,
                "H":15,
                "I":12,
                "J":12,
                "K":30
                }
    for k,v in widthDic.items():
        sheet3.column_dimensions[k].width = v 

    widthDic = {"A":15,
                "B":25,
                "C":10,
                "D":10,
                "E":10,
                "F":12,
                "G":10,
                "H":15,
                "I":12,
                "J":12,
                "K":30
                }
    for k,v in widthDic.items():
        sheet4.column_dimensions[k].width = v 

    #　高さ調節
    for row in range(3, sheet1.max_row + 1):
        sheet1.row_dimensions[row].height = 30
    sheet1.row_dimensions[1].height = 30
    sheet1.row_dimensions[2].height = 60

    for row in range(1, sheet2.max_row + 1):
        sheet2.row_dimensions[row].height = 22
    sheet2.row_dimensions[2].height = 50

    for row in range(1, sheet3.max_row + 1):
        sheet3.row_dimensions[row].height = 24

    for row in range(1, sheet4.max_row + 1):
        sheet4.row_dimensions[row].height = 24




    alignCenter = Alignment(wrapText=True,horizontal="center", vertical="center")
    sheet1 = mergeRows(sheet1, col=1, nAmount=1)
    sheet2 = mergeRows(sheet2, col=1, nAmount=1)
    sheet3 = mergeRows(sheet3, col=1, nAmount=5)
    sheet3 = mergeRows(sheet3, col=2, nAmount=5)
    sheet4 = mergeRows(sheet4, col=1, nAmount=2)
    sheet4 = mergeRows(sheet4, col=2, nAmount=2)


    r = getRowMergeInd(sheet3, nAmount=5, baseCol = 2)
    sheet3 = mergeRowsCond(sheet3, col=3, rowMergeInd=r)
    #sheet3 = setInPatientNum(sheet3, col=4, rowMergeInd=r)
    sheet3 = mergeRowsCond(sheet3, col=4, rowMergeInd=r)

    r = getRowMergeInd(sheet4, nAmount=2, baseCol = 2)
    sheet4 = mergeRowsCond(sheet4, col=3, rowMergeInd=r)
    #sheet4 = setInPatientNum(sheet4, col=4, rowMergeInd=r)
    sheet4 = mergeRowsCond(sheet4, col=4, rowMergeInd=r)



    # セル内位置
    for column in sheet1:
        for cell in column:
            if cell.column ==1:
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
            elif cell.column == 2:
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center', wrap_text = True)
            elif cell.column in (3, 4):
                cell.alignment = Alignment(horizontal = 'right', vertical = 'center', wrap_text = True)

    for row in sheet1:
        for cell in row:
            if cell.row == 2:
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center',  wrap_text = True)

    sheet1['A1'].alignment = Alignment(horizontal = 'left', vertical = 'center', wrap_text = True)


    for column in sheet2:
        for cell in column:
            if cell.column ==1:
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
            elif cell.column == 2:
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center', wrap_text = True)
            elif cell.column in (3, 4, 5, 6, 7, 8):
                cell.alignment = Alignment(horizontal = 'right', vertical = 'center', wrap_text = True)

    for row in sheet2:
        for cell in row:
            if cell.row in (1,2):
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center',  wrap_text = True)

    for column in sheet3:
        for cell in column:
            if cell.column in (1, 6, 7):
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
            elif cell.column in (2, 6, 8, 11):
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center', wrap_text = True)
            elif cell.column in (3, 4, 5, 9,10):
                cell.alignment = Alignment(horizontal = 'right', vertical = 'center', wrap_text = True)

    for row in sheet3:
        for cell in row:
            if cell.row in (2,5,6):
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center',  wrap_text = True)
            elif cell.row ==4:
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center',  wrap_text = True)

    for column in sheet4:
        for cell in column:
            if cell.column in (1, 6, 7):
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
            elif cell.column in (2, 6, 8, 11):
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center', wrap_text = True)
            elif cell.column in (3, 4, 5, 9,10):
                cell.alignment = Alignment(horizontal = 'right', vertical = 'center', wrap_text = True)

    for row in sheet4:
        for cell in row:
            if cell.row in (2,3):
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center',  wrap_text = True)
            elif cell.row ==1:
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center',  wrap_text = True)


    #文字大きさ
    for column in sheet1:
        for cell in column:
            if cell.column in (1,2,3,4):
                cell.font = Font(size=28)

    for column in sheet2:
        for cell in column:
            if cell.column in (1,2,3,4,5,6,7,8):
                cell.font = Font(size=20)

    for column in sheet3:
        for cell in column:
            if cell.column in (1,2,3,4,5,6,7,8,9,10,11):
                cell.font = Font(size=12)

    for column in sheet4:
        for cell in column:
            if cell.column in (1,2,3,4,5,6,7,8,9,10,11):
                cell.font = Font(size=12)

    #表題
    sheet1['A1'] = "【第一種・第二種感染症指定医療機関】"
    sheet1['A1'].font = Font(size=20)
    sheet3["A3"] = '新型コロナウィルス感染患者入院状況'
    sheet3['A3'].font = Font(size=20)



    #色塗り
    grayFill = PatternFill("solid", bgColor= 'bcbcbc', fgColor='bcbcbc')
    for c in "ABCD":
        sheet1[c + '2'].fill = grayFill

    for c in "ABCF":
        sheet2[c + '1'].fill = grayFill
    for c in "ABCDEFGH":
        sheet2[c + '2'].fill = grayFill

    for c in "ABCDEFGHIJK":
        sheet3[c + '5'].fill = grayFill
        sheet3[c + '6'].fill = grayFill

    for c in "ABCDE":
        sheet4[c + '2'].fill = grayFill
    for c in "EFGHIJK":
        sheet4[c + '3'].fill = grayFill


    # set colors
    redFill = PatternFill("solid",bgColor="fb0007",fgColor ="fb0007" )
    yellowFill = PatternFill("solid",bgColor="ffff87",fgColor ="ffff87" )
    whiteFill = PatternFill("solid",bgColor="ffffff",fgColor ="ffffff" )
    def conditionalColoring(ws,color,formula, rg  ):
        # differential style, conditional color formatting.
        dxf = DifferentialStyle(fill=color)
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = formula
        ws.conditional_formatting.add(rg, r)
        return(ws)

    def conditionalTwo(ws,rowBase) :
        for row in range(rowBase, ws.max_row+1):
            v1 = ws.cell(row,3)._value
            v2 = ws.cell(row,4)._value
            if v1 == None:
                v1 = 0
            if v2 == None:
                v2 = 0

            if v1 <= v2 :
                ws.cell(row,3).fill = redFill
                ws.cell(row,4).fill = redFill
            elif v1 - 1 == v2 :
                ws.cell(row,3).fill = yellowFill 
                ws.cell(row,4).fill = yellowFill

            if v1 == 0:
                if v2 == 0:
                    ws.cell(row,3).fill = whiteFill
                    ws.cell(row,4).fill = whiteFill

        return(ws)

    ws = sheet3
    sheet3 = conditionalColoring(ws = sheet3,
                color = grayFill ,
                formula=['$J7<>"" '],
                rg= f"E7:K{ws.max_row}" )


    sheet3 = conditionalColoring(ws = sheet3,
                color = whiteFill,
                formula=['$D7 = 0'],
                rg= f"C7:D{ws.max_row}")

    sheet3 = conditionalTwo(sheet3,7)

    ws = sheet4
    sheet4 = conditionalColoring(ws = sheet4,
                color = grayFill ,
                formula=['$J4<>"" '],
                rg= f"E4:K{ws.max_row}")


    sheet4 = conditionalTwo(sheet4,4)

    #罫線
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    # write in sheet1
    for row in sheet1:
        for cell in row:
            sheet1[cell.coordinate].border = border

    for row in sheet2:
       for cell in row:
            sheet2[cell.coordinate].border = border

#    for row in sheet3:
#        for cell in row:
#            sheet3[cell.coordinate].border = border

    for row in range(4, sheet3.max_row + 1):
        for col in range(1,sheet3.max_column+1):
            sheet3.cell(row=row ,column=col).border = border

    for row in sheet4:
        for cell in row:
            sheet4[cell.coordinate].border = border


    # 保存する
    book.save(output1)


def mergeTop(ws,nAmount):
    for col in range(1,ws.max_column + 1 ):
        if col in [1,2,3,4]:
            val = ws.cell(nAmount+1, col)._value
            _ = ws.cell(nAmount,col,value=val)
            ws.merge_cells(start_row= nAmount, start_column=col,
                           end_row = nAmount + 1,end_column = col )
            ws.cell(nAmount,col).alignment = baseAlign

        if col == 5:
            ws.merge_cells(start_row= nAmount, start_column=col,
                           end_row = nAmount,end_column = ws.max_column )
            _ = ws.cell(nAmount,col,value="内訳")
            ws.cell(nAmount,col).alignment = baseAlign

        ws.cell(nAmount,col).fill = darkGreyFill
    return(ws)

def conditionalColoring(ws):
    # differential style, conditional color formatting.
    dxf = DifferentialStyle(fill=darkGreyFill)
    r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    r.formula = ['$J4<>""']
    ws.conditional_formatting.add(f"E4:K{ws.max_row}", r)
    return(ws)

def mergeRows(ws,col= 1,nAmount = 3 ):
    '''二次医療圏と医療機関名の縦方向をマージする用の関数です。
    '''
    macV = ws.cell(nAmount+2,col)._value
    macRow = nAmount+2

    for row in range(nAmount + 2, ws.max_row + 1):
        colV = ws.cell(row,col)._value
        if macV != colV:
            ws.merge_cells(start_row= macRow, start_column=col,
                           end_row = row-1,end_column = col )
            ws.cell(macRow,col).alignment = alignCenter
            macRow = row
            macV = colV

    return(ws)

def getMergedCellType():
    wb_ = Workbook()
    ws_ = wb_.active
    ws_.merge_cells('A1:A4')
    tp_ = type(ws_['A2'])
    del wb_
    return(tp_)

def getRowMergeInd(ws,nAmount,baseCol=2):
    '''医療機関名からmergeの基準となるrowのindexを取得する関数です。
    '''
    mergedCellType = getMergedCellType()
    r = []
    for row in range(nAmount + 2, ws.max_row + 1):
        val = ws.cell(row,baseCol)
        if not isinstance(val, mergedCellType):
            r.append(row)
    return(r)

def mergeRowsCond(ws,col,rowMergeInd):
    for i in range(len(rowMergeInd) -1 ):
        row1 = rowMergeInd[i]
        row2 = rowMergeInd[i+1]
        ws.merge_cells(start_row= row1, start_column=col,
                               end_row = row2-1,end_column = col )
        ws.cell(row1,col).alignment = alignCenter
    return(ws)

def setInPatientNum(ws,col,rowMergeInd):
    for i in range(len(rowMergeInd) -1 ):
        row1 = rowMergeInd[i]
        row2 = rowMergeInd[i+1]
        sum_ = 0
        for j in range(row1,row2):
            v_ = ws.cell(j,col)._value
            if isinstance(v_,int) :
                sum_ += v_
            # TO DO:
            # if value which is not "int" is  inserted, how to behave
        _ =  ws.cell(row1,col,value=sum_)
    return(ws)


if __name__ == "__main__":
    main(pL.newStylePatient, pL.formattedHospital)



