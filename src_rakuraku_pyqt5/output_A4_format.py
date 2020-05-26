import pandas as pd
import numpy as np
import datetime
import os
import sys
import _utils

import pathList as pL


if os.name == "nt":
    import locale
    locale.setlocale(locale.LC_CTYPE, "Japanese_Japan.932")

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color,colors,  Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

program = "output_A4_formatj.py"



def main(path):
    try:
        df = pd.read_excel(path, header = 1,
                        sheet_name = pL.patientSheetName,
                        encoding="cp932") 
        dataProcessing(df,path)

        _utils.createErrorCheckFile(path,program = program) 
    except:
        return("読み取りエラーです")

    return(False)



def dataProcessing(df,path):
    df = df.replace({"(.*)\xa0":r"\1"},regex=True)
    df = df.replace({"(.*)\n":r"\1"},regex=True) 

    inDate = "入院日"
    outDate = "退院日"
    onset  = "発症日"
    df[inDate]  = _utils.convertDatetime(df,inDate) 
    df[outDate] = _utils.convertDatetime(df,outDate) 
    df[onset]   = _utils.convertDatetime(df,onset ) 

    df[inDate] = pd.to_datetime(df[inDate].astype(str), format='%Y-%m-%d')
    df[outDate] = pd.to_datetime(df[outDate].astype(str), format='%Y-%m-%d')
    df[onset] = pd.to_datetime(df[onset].astype(str), format='%Y-%m-%d')

    job2 = '職業２'
    jobCom = "職業：備考"
    df[job2] = df[job2].replace(np.nan,"")
    df[jobCom] = df[jobCom].replace(np.nan,"")
    df[job2] = df[job2] + "," + df[jobCom] 
    df[job2] = df[job2].replace(',','')

    df1 = df[pL.A4OutputCols].copy()

    underScoreList  =  ['保健所', '性別','国籍', '職業分類',  '濃厚接触者の観察状況（公表）',
            '入院医療機関（現在）','身体状況（現在の症状）' ,'入院病床（現在）', '治療機器' ]

    for c in underScoreList:
        df1[c] = df1[c].replace({".+_(.+)":r"\1"},regex=True)


    df1 = df1.rename(columns=pL.repOutput) 


    for i, a in zip(df1.index, df1['氏名'].values):
        var = name(str(a))
        df1.loc[i,'氏名'] = var
        
    #df1['氏名'].value_counts()
    df2 = df1.copy()

    dir_ = _utils.getOutputDir()
    pathOutput = dir_ + path.split("/")[-1][:-5] + "_A4format.xlsx"

    with pd.ExcelWriter(pathOutput, engine = 'openpyxl' , mode='wa', datetime_format='mm/dd') as writer: 
        df2.to_excel(writer, sheet_name='最新', index=False)


    excel_formatting(pathOutput)

def excel_formatting(pathOutput):
    # ブックを取得
    book = load_workbook(pathOutput)
    # シートを取得
    sheet = book['最新']


    #行挿入
    SheetRowNo=2
    sheet.insert_rows(SheetRowNo)

    #太字
    font = Font(bold=True)

    #セル結合
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:B2')
    sheet.merge_cells('C1:C2')
    sheet.merge_cells('D1:D2')
    sheet.merge_cells('E1:E2')
    sheet.merge_cells('F1:F2')
    sheet.merge_cells('G1:G2')
    sheet.merge_cells('H1:H2')
    sheet.merge_cells('I1:I2')
    sheet.merge_cells('J1:J2')
    sheet.merge_cells('K1:K2')
    sheet.merge_cells('L1:L2')
    sheet.merge_cells('M1:M2')
    sheet.merge_cells('N1:N2')
    sheet.merge_cells('O1:O2')
    sheet.merge_cells('P1:P2')

    sheet.merge_cells('Q1:X1')
    sheet["Q1"] = "陰性検査（検査日、結果）"

    sheet.merge_cells('Q2:R2')
    sheet["Q2"] = "一回目"
    sheet["Q2"].font = font

    sheet.merge_cells('S2:T2')
    sheet["S2"] = "二回目"
    sheet["S2"].font = font

    sheet.merge_cells('U2:V2')
    sheet["U2"] = "三回目"
    sheet["Q2"].font = font

    sheet.merge_cells('W2:X2')
    sheet["W2"] = "四回目"
    sheet["W2"].font = font


    sheet.merge_cells('Y1:Y2')
    sheet.merge_cells('Z1:Z2')
    sheet.merge_cells('AA1:AA2')

    # 行の高さを変更
    #sheet.row_dimensions[1].height = 30
    # 列の幅を変更
    sheet.column_dimensions['a'].width = 3.83
    sheet.column_dimensions['b'].width = 9.83
    sheet.column_dimensions['c'].width = 6.33
    sheet.column_dimensions['d'].width = 10
    sheet.column_dimensions['e'].width = 7.5
    sheet.column_dimensions['f'].width = 6.33
    sheet.column_dimensions['g'].width = 12
    sheet.column_dimensions['h'].width = 14
    sheet.column_dimensions['i'].width = 20
    sheet.column_dimensions['j'].width = 14
    sheet.column_dimensions['k'].width = 14
    sheet.column_dimensions['l'].width = 14
    sheet.column_dimensions['m'].width = 14
    sheet.column_dimensions['n'].width = 20
    sheet.column_dimensions['o'].width = 9
    sheet.column_dimensions['p'].width = 9
    sheet.column_dimensions['q'].width = 6
    sheet.column_dimensions['r'].width = 8
    sheet.column_dimensions['s'].width = 6
    sheet.column_dimensions['t'].width = 8
    sheet.column_dimensions['u'].width = 6
    sheet.column_dimensions['v'].width = 8
    sheet.column_dimensions['w'].width = 6
    sheet.column_dimensions['x'].width = 8
    sheet.column_dimensions['y'].width = 9
    sheet.column_dimensions['z'].width = 12
    sheet.column_dimensions['aa'].width = 9
    sheet.column_dimensions['ab'].width = 9

    #　高さ調節
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 29

    # セル内位置
    for column in sheet:
        for cell in column:

            if cell.column in (1,2, 4, 5, 6, 15,16,17,18,19,20,21,22,23,24,25):
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)

            elif cell.column in (3, 7, 8, 9, 10, 11, 12, 13, 14, 26, 27, 28):
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center', wrap_text = True)

    for row in sheet:
        for cell in row:

            if cell.row == 1:
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center',  wrap_text = True)

    #罫線
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    # write in sheet
    for row in sheet:
        for cell in row:
            sheet[cell.coordinate].border = border

    #色塗り（カラム名）
    greenFill = PatternFill("solid", fgColor='15a535')
    sheet.cell(1,1).fill = greenFill
    sheet.cell(1,2).fill = greenFill
    sheet.cell(1,4).fill = greenFill
    sheet.cell(1,5).fill = greenFill
    sheet.cell(1,8).fill = greenFill
    sheet.cell(1,10).fill = greenFill
    sheet.cell(1,11).fill = greenFill


    # set colors
    redFill = PatternFill("solid",bgColor="fb0007",fgColor ="fb0007" )
    def conditionalColoring(ws):
        # differential style, conditional color formatting.
        dxf = DifferentialStyle(fill=redFill)
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = ['$Y3 = "死亡"']
        ws.conditional_formatting.add(f"A3:AA{ws.max_row}", r)
        return(ws)
    conditionalColoring(sheet)

    yellowFill = PatternFill("solid",bgColor="ffff87",fgColor ="ffff87" )
    def conditionalColoring(ws):
        # differential style, conditional color formatting.
        dxf = DifferentialStyle(fill=yellowFill)
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = ['$Y3 = "軽・中等症"']
        ws.conditional_formatting.add(f"A3:AA{ws.max_row}", r)
        return(ws)
    conditionalColoring(sheet)

    blueFill = PatternFill("solid",bgColor="d5e6f3",fgColor ="d5e6f3" )
    def conditionalColoring(ws):
        # differential style, conditional color formatting.
        dxf = DifferentialStyle(fill=blueFill)
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = ['$Y3 = "陰性確認"']
        ws.conditional_formatting.add(f"A3:AA{ws.max_row}", r)
        return(ws)
    conditionalColoring(sheet)

    pinkFill = PatternFill("solid",bgColor="f5c3c1",fgColor ="f5c3c1" )
    def conditionalColoring(ws):
        # differential style, conditional color formatting.
        dxf = DifferentialStyle(fill=pinkFill)
        r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        r.formula = ['$Y3 = "重症"']
        ws.conditional_formatting.add(f"A3:AA{ws.max_row}", r)
        return(ws)
    conditionalColoring(sheet)


    sheet.freeze_panes = 'L3'
    sheet.auto_filter.ref = 'A2:AA500' # cf. ws.dimensions, add auto_filter.

    # 保存する

    book.save(pathOutput)




def daterule(x):
    if x != x:
        day = np.nan
    try:
        if 40000 < int(x) < 50000:
            day = datetime.date(1899, 12, 31) + datetime.timedelta(days = int(x) - 1)
    except:
        pass

    try:
        day = x.date()
    except:
#        print(type(x),x)
        day = x
    return(day)
        
def name(x):
    try:
#        print(type(x))
        if '(' in str(x):
            var = x.replace('(' , '\n(')
        elif '（' in str(x):
            var = x.replace('（' , '\n（')
        elif x == 'nan':
            var = ''
        else:
            var = str(x)
        
        var = var.replace('\n\n(', '\n(')
        var = var.replace('\n\n（', '\n（')
        
        return(var)
    except:
        print('exception', x)
 
if __name__ == "__main__":
    import pathList as pL
    main(pL.newStylePatient)
