# TO DO : 
# date is odd, please change for good format. 

import pandas as pd
import numpy as np
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color,colors,  Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


# define colors 
yellowFill = PatternFill("solid", fgColor=colors.YELLOW)
lightSteelBlue = PatternFill("solid", fgColor="d8e3ee")
beige = PatternFill("solid", fgColor="e9eedb")

# define variables
phcCol = "保健所"
ageCol = "年代"
sexCol = "性別"
healthInstCol = "入院医療機関（現在）"

def main(path,shName):
    pullTab = pd.read_excel(path,sheet_name="プルタブシート",header=1)
    wb = load_workbook(filename = path)

    ws = wb[shName]
    ws.insert_rows(0)

    for col in range(1, ws.max_column+1):
        _ = ws.cell(row=1,column=col,value = col)
        if col in [3,41,42]:
            ws.cell(1,col).fill = lightSteelBlue
            ws.cell(2,col).fill = lightSteelBlue
        elif col in [30,31,32,33,34,35,36,37,38,39,40,43,45,46]:
            ws.cell(1,col).fill = beige
            ws.cell(2,col).fill = beige
        else:
            ws.cell(1,col).fill = yellowFill
            ws.cell(2,col).fill = yellowFill
        
        ws.cell(1,col).alignment = Alignment(wrapText=True)
        ws.cell(2,col).alignment = Alignment(wrapText=True)
    
    ws.freeze_panes = 'O3'
    ws.row_dimensions[2].height = 40
    ws.auto_filter.ref = 'A2:AX160' # cf. ws.dimensions, add auto_filter. 

    # data validation / create drop down for all rows.     
    ws = setDropDown(ws,pullTab,phcCol,"D")
    ws = setDropDown(ws,pullTab,ageCol,"E")
    ws = setDropDown(ws,colStr="F",fm1='"公表,非公表"')
    ws = setDropDown(ws,colStr="H",fm1='"男,女,不明"')
    ws = setDropDown(ws,colStr="I",fm1='"公表,非公表"')

    wb.save(path)


def replaceSheet(path1,sheetName1, path2,sheetName2 = None,pathOutput=None):
    if pathOutput == None:
        pathOutput = path1[:-5] + "_output.xlsx"

    wb = load_workbook(filename = path1)
    wb.remove(wb[sheetName1])
    wb.create_sheet(index=0,title=sheetName1)
    ws1 = wb[sheetName1]

    wbHidden = load_workbook(filename = path2)

    if sheetName2 == None:
        ws2 = wbHidden.active
    else:
        ws2 = wbHidden[sheetName2]

    for row in ws2:
        for cell in row:
            ws1[cell.coordinate].value = cell.value
    wb.save(pathOutput)


    
def setDropDown(ws,pullTab=None,colName=None,colStr=None,fm1=None): 
    if fm1 == None:
        s = ",".join(pullTab[colName].dropna().unique())
        fm1 =  f'"{s}"'
    dv = DataValidation(type="list", formula1=fm1)
    dv.ranges = f'{colStr}3:{colStr}1048576'
    ws.add_data_validation(dv)

        
    return(ws)
