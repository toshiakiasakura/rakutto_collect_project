# ======================================================================
# Project Name    : rakutto_collect_project
# File Name       : merge_positive.py
# Encoding        : utf-8
# Copyright © 2020 Toshiaki Asakura. All rights reserved.
# ======================================================================

import pandas as pd
import numpy as np
import datetime
import _utils
import os 
import pathList as pL

from openpyxl import Workbook
from openpyxl import load_workbook

if os.name == "nt":
    import locale
    locale.setlocale(locale.LC_CTYPE, "Japanese_Japan.932")

program = "merge_positive.py"

res = "結果"
testType = "検査種別"
infNum = "感染事例番号_陰性確認検査含む"
phc = "担当\n保健所"
phcCol = "保健所"
testDate = "検査実施日"
negTimes = "陰性確認検査の回数"
colList = [res,testType,infNum,phc,testDate,negTimes]

difDays = 1

def main(path1, path2):
    df =pd.read_excel(path1,sheet_name=pL.patientSheetName ,
            encoding="cp932",header = 1)
    test = pd.read_excel(path2, header = 1,  encodeing="cp932" ,sheet_name="Sheet1")

    mergeTest(df,test,path1)
    _utils.createErrorCheckFile(path1,program = program) 
    _utils.createErrorCheckFile(path2,program = program) 
    return(False)


def mergeTest(df,test,path1):
    shName = pL.patientSheetName
    merged = createPosDataFrame(df,test)
    negStSer , negPassSer = createNegSeries(df,test)

    # replace values 
    dir_ = _utils.getOutputDir()
    pathOutput = dir_ +  path1.split("/")[-1][:-5] + "_rep.xlsx"
    wb = load_workbook(filename = path1)
    ws = wb[shName]

    colIndex = _utils.getColIndex(ws,"陽性確定日")
    replaceRowValues(ws,series_=merged    ,col=colIndex)

#    colIndex = _utils.getColIndex(ws,"陰性結果開始日")
#    replaceRowValues(ws,series_=negStSer  ,col=colIndex,condLis=pL.testFilter)
#    colIndex = _utils.getColIndex(ws,"陰性結果確認日")
#    replaceRowValues(ws,series_=negPassSer,col=colIndex,condLis=pL.testFilter)

    wb.save(pathOutput)

def createPosDataFrame(df,test):
    # merge part 
    test = test[colList]
    if test[testDate].dtype == "datetime64[ns]" : 
        test[testDate] = test[testDate].dt.date
    cond1 = test[res] == '陽性'
    cond2 = test[testType] == '診断'
    cond3 = test[infNum].isnull() == False

    testFil = test.loc[cond1&cond2&cond3]

    # 診断結果が陽性かつ感染事例番号が振られている人について
    check = testFil[infNum].value_counts() > 1 
    if check.sum() > 0 :
        raise Exception(f"'{infNum}' が二回書かれている人がいます。" )
    # 日付がある人だけ抜き出し
    testFil = testFil.dropna(subset = ["検査実施日"])
    numDateDF = testFil[[infNum,testDate]]


    pos = "陽性確定日"
    merged = pd.merge(df,numDateDF,left_on = '№',right_on=infNum,how="left")
    print(merged[testDate].dtypes) 
    print(merged[testDate].unique()) 
    merged = merged[testDate] + datetime.timedelta(days = difDays) 

    return(merged)

def createNegSeries(df,test):
    # timestamp handling
    test[testDate] = test[testDate].dt.date

    # get unique inf. number.
    uni = test[infNum].unique()
    cond1 = uni == uni
    cond2 = uni != -99
    uni = uni[cond1&cond2]

    # make dataframe.
    # 前提1. 陽性が出てから感染事例番号が振られる。つまり、その前に何回診断検査が陰性になったかは分からない。
    # 前提2. 診断検査が2回出てくる場合は考えない。
    # 前提3. 陰性確認検査通過日は、初めてpassした日を抜き出す。　
    No = "№"
    negSt = "隂開始日"
    negPass = "陰通過日"
    dicMerge = {No:[],negSt:[],negPass:[]}

    for u in uni:
        cond = test[infNum] == u
        testM = test.loc[cond].sort_values(by=testDate)
        dicMerge[No].append(u)

        cond = testM[testType] == "陰性確認"
        testMM = testM.loc[cond]
        if testMM.shape[0] > 0:
            #display(testM)
            v = testMM.iloc[0][testDate]
        else:
            v = np.nan
        dicMerge[negSt].append(v)

        cond1  = testM[negTimes] == 2
        cond2 = testM[res] == "（－）"
        testMM = testM.loc[cond1&cond2]
        if testMM.shape[0] > 0:
            v = testMM.iloc[0][testDate]
        else:
            v = np.nan
        dicMerge[negPass].append(v)

    testDF = pd.DataFrame(dicMerge)
    merged = pd.merge(df,testDF,left_on = '№',right_on=No,how="left")

    negStSer = merged[negSt] + datetime.timedelta(days = difDays) 
    negPassSer = merged[negPass] + datetime.timedelta(days = difDays) 

    #print(merged[negSt].value_counts() )
    #print(negStSer.value_counts() ) 

    return(negStSer,negPassSer)


def replaceRowValues(ws, series_,col,condLis=None):
    vals = series_.values

    for row in range(series_.shape[0]):
        if condLis == None:
            _ = ws.cell(row=row+3 ,column=col,value = vals[row])
        else:
            # 保健所ごと
            hpcIndex = _utils.getColIndex(ws,"保健所")
            v = ws.cell(row=row+3, column= hpcIndex)._value 
            if v not in  condLis:
                #print(v)
                _ = ws.cell(row=row+3, column=col, value = vals[row])

    return(ws) 


if __name__ == "__main__":
    import pathList as pL
    main(pL.newStylePatient, pL.testData) 
