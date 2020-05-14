# ======================================================================
# Project Name    : rakutto_collect_project
# File Name       : in_hosp_status.py
# Encoding        : utf-8
# Copyright © 2020 Toshiaki Asakura. All rights reserved.
# ======================================================================

import numpy as np
import pandas as pd
import datetime
import openpyxl
import os

import _utils
import pathList as pL

import warnings
warnings.simplefilter('ignore')

if os.name == "nt":
    import locale
    locale.setlocale(locale.LC_CTYPE, "Japanese_Japan.932")

program = "simple_in_hosp_status.py"


# for global variables of formatExcelStyle() 
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color,colors,Border,Side,  Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

# set colors
darkGreyFill = PatternFill("solid",bgColor="A9A9A9",fgColor ="A9A9A9" )
baseAlign = Alignment(wrapText=True,horizontal="center", vertical="top")
alignCenter = Alignment(wrapText=True,horizontal="center", vertical="center")
borderType = 'thin'
border = Border(left=Side(border_style=borderType,
                           ),
                 right=Side(border_style=borderType,
                           ),
                 top=Side(border_style=borderType,
                          ),
                 bottom=Side(border_style=borderType,
                           ),
                 outline=Side(border_style=borderType),
               )
#borderType = 'thick'
#borderThick = Border(left=Side(border_style=borderType,
#                           ),
#                 right=Side(border_style=borderType,
#                           ),
#                 top=Side(border_style=borderType,
#                          ),
#                 bottom=Side(border_style=borderType,
#                           ),
#                 outline=Side(border_style=borderType),
#               )

hosp = "入院医療機関（現在）"
sex = "性別"
address = "居住地"
bed = "入院病床（現在）"
No = "№"
out = "退院日"
physical = "身体状況（現在の症状）"

numHosp = "入院患者数"
sh1 = "第一種・第二種感染症指定医療機関_入院状況"
sh2 = "その他医療機関_入院状況"
sh3 = "(入院者のみ)第一種・第二種感染症指定医療機関_入院状況"
sh4 = "(入院者のみ)その他医療機関_入院状況"
shNames = [sh1,sh2,sh3,sh4]


def main(path ):
    try:
        df = pd.read_excel(path,
                        sheet_name=pL.patientSheetName,
                        header = 1,
                        encodeing="cp932"
                        )
    except:
        errorMsg = "データが読み込めませんでした。\n"
        import traceback
        errorMsg  +=  traceback.format_exc()
        print(errorMsg)
        return(errorMsg)


    errorMsg = getErrorMsg(df)
    if errorMsg == False:
        try:
            conversion(df,path)
            formatExcelStyle(path)
            _utils.createErrorCheckFile(path,program = program) 
        except:
            errorMsg  = "予期せぬエラーが発生しました。\n"
            import traceback
            errorMsg  +=  traceback.format_exc()
            print(errorMsg)


    print(errorMsg)    
    return(errorMsg)


def getErrorMsg(df):
    return(False)


def conversion(df,path):
    df = df.dropna(axis=0,how="all")

    com  = '備考'
    df[com] = ""
    for c in pL.commentCols: 
        df[com] += " " + df[c].replace(np.nan,'').astype(str) 


    df1 = df[[hosp, No, '年代', 
            sex , '居住地' , '入院日', '退院日',bed ,physical ,'備考']]

    df1[physical]  = df1[physical].replace({".+_(.+)":r"\1"}, regex=True) 

    df1[numHosp] = 0
    condOut = df1[out].dt.date.apply(type) == type(datetime.date(2000,1,12))
    df1.loc[1^condOut,numHosp] = 1 

    df1 = df1.sort_values(by=[hosp,No])
    cols = [hosp ,numHosp, '№', '年代', '性別', 
            '居住地', '入院日', '退院日', bed ,physical , '備考']
    df1 = df1[cols]

    condInf = df1[bed] == "01_感染症病床"
    
    for c in [sex,bed]:
        df1[c] = df1[c].replace({".+_(.+)":r"\1"},regex=True)

    dfInf = df1.loc[condInf]
    dfOther = df1.loc[1^condInf]
    

    dir_ = _utils.getOutputDir()
    pathOutput = dir_ + path.split("/")[-1][:-5] + f"_入院状況.xlsx"
    with pd.ExcelWriter(pathOutput , engine="openpyxl", mode="wa",
            datetime_format='mm/dd') as writer:
        dfInf.to_excel(writer,index=False ,
                sheet_name=sh1)
        dfOther.to_excel(writer,index=False,
                sheet_name=sh2)

        dfInfIn = dfInf[ dfInf[numHosp] == 1 ] 
        dfInfIn.to_excel(writer,index=False ,
                sheet_name=sh3)

        dfOtherIn = dfOther[ dfOther[numHosp] == 1 ] 
        dfOtherIn.to_excel(writer,index=False ,
                sheet_name=sh4)


def formatExcelStyle(path):
    dir_ = _utils.getOutputDir()
    pathOutput = dir_ + path.split("/")[-1][:-5] + f"_入院状況.xlsx"
    # ブックを取得
    book = openpyxl.load_workbook(pathOutput)

    # シートを取得
    for sh in shNames:
        ws = book[sh]
        formatWS(ws)

    # 保存する
    book.save(pathOutput)

def formatWS(ws):

    #行挿入
    ws.insert_rows(1)
    ws.insert_rows(2)
    ws.insert_rows(3)
    ws.insert_rows(4)
    ws.insert_rows(5)


    #セル結合
    ws.merge_cells('A5:A6')
    ws["A5"] = "医療機関名"


    ws.merge_cells('B5:B6')
    ws["B5"] = '入院\n患者数'

    ws.merge_cells('C5:J5')
    ws["C5"] = "内  訳"

    colVals = {
            "C": 'No.', 
            "D": '年齢',
            "E": '性別',
            "F": '居住地',
            "G": '入院月日',
            "H": '退院月日',
            "I": '病床',
            "J": '身体状況',
            "K": '備考'}
    for k,v in colVals.items():
        ws[k + "6"] = v

    ws['K2'] = '取扱注意'

    ws.merge_cells('A3:K3')
    ws.merge_cells('B4:I4')
    ws['B4'] = '【第一種・第二種感染症指定医療機関】'


    #列幅

    widthDic = {"A":50,
                "B":10,
                "C":19,
                "D":12,
                "E":10,
                "F":15,
                "G":12,
                "H":12,
                "I":15,
                "J":15,
                "K":20
                }
    for k,v in widthDic.items():
        ws.column_dimensions[k].width = v 


    #　高さ調節

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 24



    alignCenter = Alignment(wrapText=True,horizontal="center", vertical="center")
    ws = mergeRows(ws, col=1, nAmount=5)


    r = getRowMergeInd(ws, nAmount=5, baseCol = 1)
    ws = setInPatientNum(ws, col=2, rowMergeInd=r)
    ws = mergeRowsCond(ws, col=2, rowMergeInd=r)



    for column in ws:
        for cell in column:
            if cell.column in (1, 4, 5):
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
            elif cell.column in ( 6,):
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center', wrap_text = True)
            elif cell.column in ( 3, 7, 8, 9, 10, 11):
                cell.alignment = Alignment(horizontal = 'right', vertical = 'center', wrap_text = True)

    for row in ws:
        for cell in row:
            if cell.row in (2,5,6):
                cell.alignment = Alignment(horizontal = 'center', vertical = 'center',  wrap_text = True)
            elif cell.row ==4:
                cell.alignment = Alignment(horizontal = 'left', vertical = 'center',  wrap_text = True)

    #文字大きさ
    for column in ws:
        for cell in column:
            if cell.column in (1,2,3,4,5,6,7,8,9,10,11):
                cell.font = Font(size=12)

    #表題
    ws["A3"] = '新型コロナウィルス感染患者入院状況'
    ws['A3'].font = Font(size=20)


    #色塗り
    grayFill = PatternFill("solid", bgColor= 'bcbcbc', fgColor='bcbcbc')

    for c in "ABCDEFGHIJK":
        ws[c + '5'].fill = grayFill
        ws[c + '6'].fill = grayFill


    # set colors
    redFill = PatternFill("solid",bgColor="fb0007",fgColor ="fb0007" )
    yellowFill = PatternFill("solid",bgColor="ffff87",fgColor ="ffff87" )
    whiteFill = PatternFill("solid",bgColor="ffffff",fgColor ="ffffff" )

    ws = ws
    ws = conditionalColoring(ws = ws,
                color = grayFill ,
                formula=['$H7<>"" '],
                rg= f"C7:K{ws.max_row}" )

    #罫線
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
#    for row in ws:
#        for cell in row:
#            ws[cell.coordinate].border = border

    for row in range(3, ws.max_row + 1):
        for col in range(1,ws.max_column+1):
            ws.cell(row=row ,column=col).border = border



def conditionalColoring(ws,color,formula, rg  ):
    # differential style, conditional color formatting.
    dxf = DifferentialStyle(fill=color)
    r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
    r.formula = formula
    ws.conditional_formatting.add(rg, r)
    return(ws)

def mergeRows(ws,col= 1,nAmount = 3 ):
    '''二次医療圏と医療機関名の縦方向をマージする用の関数です。
    '''
    macV = ws.cell(nAmount+2,col)._value
    macRow = nAmount+2

    for row in range(nAmount + 2, ws.max_row + 2):
        colV = ws.cell(row,col)._value
        if macV != colV:
            ws.merge_cells(start_row= macRow, start_column=col,
                           end_row = row-1,end_column = col )
            ws.cell(macRow,col).alignment = alignCenter
            macRow = row
            macV = colV

    return(ws)

def getMergedCellType():
    wb_ = Workbook()
    ws_ = wb_.active
    ws_.merge_cells('A1:A4')
    tp_ = type(ws_['A2'])
    del wb_
    return(tp_)

def getRowMergeInd(ws,nAmount,baseCol=2):
    '''医療機関名からmergeの基準となるrowのindexを取得する関数です。
    '''
    mergedCellType = getMergedCellType()
    r = []
    for row in range(nAmount + 2, ws.max_row + 1):
        val = ws.cell(row,baseCol)
        if not isinstance(val, mergedCellType):
            r.append(row)
    return(r)

def mergeRowsCond(ws,col,rowMergeInd):
    for i in range(len(rowMergeInd) -1 ):
        row1 = rowMergeInd[i]
        row2 = rowMergeInd[i+1]
        ws.merge_cells(start_row= row1, start_column=col,
                               end_row = row2-1,end_column = col )
        ws.cell(row1,col).alignment = alignCenter
    return(ws)

def setInPatientNum(ws,col,rowMergeInd):
    for i in range(len(rowMergeInd) -1 ):
        row1 = rowMergeInd[i]
        row2 = rowMergeInd[i+1]
        sum_ = 0
        for j in range(row1,row2):
            v_ = ws.cell(j,col)._value
            if isinstance(v_,int) :
                sum_ += v_
            # TO DO:
            # if value which is not "int" is  inserted, how to behave
        _ =  ws.cell(row1,col,value=sum_)
    return(ws)


if __name__ == "__main__":
    main(pL.newStylePatient)



